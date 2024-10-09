
#https://python-charts.com/es/colores/#google_vignette
#https://davapps.com/loom-video-downloader/#url=https://www.loom.com/share/99ad17358c3541a6bca696ed16ae125a?sid=da7bfeef-daeb-4d4a-8b68-fbbb43a1b6e8
#https://www.flaticon.es/iconos-gratis/lupa
#https://emojitool.com/
#https://www.delftstack.com/es/tutorial/tkinter-tutorial/tkinter-menubar/
#https://tutorialesprogramacionya.com/pythonya/detalleconcepto.php?codigo=65
#https://blog.furas.pl/python-tkinter-pandastable-examples-gb.html
#https://es.stackoverflow.com/questions/340278/como-mostrar-un-dataframe-en-tkinter
#https://www.delftstack.com/es/tutorial/tkinter-tutorial/tkinter-menubar/
#https://www.youtube.com/watch?v=or2ibHcZkSY
from tkinter import *
from tkinter import ttk
import tempfile
import win32api
import win32print
import win32ui
import win32con
import string
import datetime
import matplotlib.pyplot as plt
import numpy as np
from sklearn.datasets import load_iris
import pandas as pd
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
import tkinter as tk
from tkinter import filedialog
from conexionInventario import ComunicacionInventario
import mysql.connector as sql 
import pathlib
from pathlib import Path
import os
import sqlite3
from sqlite3 import Error
import datetime
import xlsxwriter
from collections import abc
import csv
from tkinter import ttk
import tkinter
import sys
import tkinter.font as tkFont
from tkinter import font 
import openpyxl
from tkinter import messagebox as mb
from tkinter import messagebox
import tkinter.messagebox as tkmb
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl import Workbook
import json
import mysql.connector  #pip install mysql-connector-python
import tkinter as tk
from pandastable import Table, TableModel
import pymysql
from tkinter import *
from tkinter import messagebox
import mysql.connector as sql
import pandas as pd
from pandastable import Table
from tabulate import tabulate
import numpy as np
import tkinter as tk


class DataFrameTable(tk.Frame):
    def __init__(self, parent=None, df=pd.DataFrame()):
        super().__init__()
        self.parent = parent
        self.table = Table(
            self, dataframe=df,
            showtoolbar=False,
            showstatusbar=True,
            editable=False)
        self.table.show()


class Aplicacion(tk.Frame):
    def __init__(self,parent=None, df=pd.DataFrame()):
        self.ventana1=tk.Tk()
        self.ventana1.title("MODULO ADMINISTRATIVO")
        self.ventana1.configure(background='#FFFFFF')#'#00BFBF'
        self.ventana1.geometry('1550x640')
        self.ventana1.iconbitmap(sys.executable)
        self.base_datos_inventario = ComunicacionInventario()
        self.portada()
        menubar1 = tk.Menu(self.ventana1)
        self.ventana1.config(menu=menubar1)
        self.labelframe1=ttk.LabelFrame(self.ventana1)#PORTADAS
        self.labelframe15=ttk.LabelFrame(self.ventana1)#DATOSDELA EMPREESA
        self.labelframe16=ttk.LabelFrame(self.ventana1)#MODULO DE COMPRAS
        self.labelframe17=ttk.LabelFrame(self.ventana1)#CODIFICACION DE PRODCUTOS
        self.labelframe18=ttk.LabelFrame(self.ventana1)#MODULO DE SALIDA
        self.labelframe19=ttk.LabelFrame(self.ventana1)#REPORTE INVENTARIOS ALL Items
        self.labelframe20=ttk.LabelFrame(self.ventana1)#REPORTE INVENTARIOS ONE Items
        self.labelframe21=ttk.LabelFrame(self.ventana1)#REPORTE INVENTARIOS INVENTARIO INCIAL
        self.labelframe22=ttk.LabelFrame(self.ventana1)#ELIMINAR ITEM MODULO DE COMPRAS
        self.labelframe23=ttk.LabelFrame(self.ventana1)#ELIMINAR ITEM MODULO DE SALIDAS
        opciones1 = tk.Menu(menubar1)
        opciones1.add_command(label="Ingrese Datos Fiscales de La Empresa", command=lambda:[self.datos_de_la_empresa(),self.labelframe1.destroy(),self.labelframe16.destroy(),self.labelframe18.destroy(),self.labelframe19.destroy(),self.labelframe20.destroy(),self.labelframe21.destroy(),self.labelframe22.destroy(),self.labelframe23.destroy()])#self.labelframe17
        menubar1.add_cascade(label="Datos de la Empresa", menu=opciones1) 
        opciones2 = tk.Menu(menubar1)
        opciones2.add_command(label="Ingresar Codificacion de Productos", command=lambda:[self.ingresar_codigo(),self.labelframe1.destroy(),self.labelframe16.destroy(),self.labelframe18.destroy(),self.labelframe19.destroy(),self.labelframe20.destroy(),self.labelframe21.destroy(),self.labelframe22.destroy(),self.labelframe23.destroy()])#self.labelframe17
        menubar1.add_cascade(label="Codificación", menu=opciones2)  
        opciones3 = tk.Menu(menubar1)
        opciones3.add_command(label="Cargar Inv. Inicial", command=lambda:[self.widgets_carga_inventario_inicial(),self.labelframe1.destroy(),self.labelframe15.destroy(),self.labelframe16.destroy(),self.labelframe17.destroy(),self.labelframe18.destroy(),self.labelframe19.destroy(),self.labelframe20.destroy(),self.labelframe22.destroy(),self.labelframe23.destroy()])#self.labelframe21
        menubar1.add_cascade(label="Inventario Inicial", menu=opciones3)  
        opciones4 = tk.Menu(menubar1)
        opciones4.add_command(label="Compras", command=lambda:[self.widgets_modulos_compras(),self.labelframe1.destroy(),self.labelframe15.destroy(),self.labelframe17.destroy(),self.labelframe18.destroy(),self.labelframe19.destroy(),self.labelframe20.destroy(),self.labelframe21.destroy(),self.labelframe22.destroy(),self.labelframe23.destroy()])#self.labelframe16
        menubar1.add_cascade(label="Modulos de Compras", menu=opciones4) 
        opciones5 = tk.Menu(menubar1)
        opciones5.add_command(label="Salidas", command=lambda:[self.widgets_modulos_salidas(),self.labelframe1.destroy(),self.labelframe15.destroy(),self.labelframe16.destroy(),self.labelframe17.destroy(),self.labelframe19.destroy(),self.labelframe20.destroy()])#self.labelframe18
        menubar1.add_cascade(label="Modulos de Salidas", menu=opciones5) 
        opciones6 = tk.Menu(menubar1)
        opciones6.add_command(label="Inventario Final", command=lambda:[self.widgets_reportes_all_inventarios(),self.widgets_reportes_all_inventario1(),self.labelframe16.destroy(),self.labelframe17.destroy(),self.labelframe18.destroy(),self.labelframe20.destroy(),self.labelframe21.destroy(),self.labelframe22.destroy(),self.labelframe23.destroy()])#self.labelframe19
        opciones6.add_command(label="Inventario Final/Producto", command=lambda:[self.widgets_reportes_one_inventarios(),self.widgets_reportes_all_inventario2(),self.labelframe16.destroy(),self.labelframe17.destroy(),self.labelframe18.destroy(),self.labelframe19.destroy(),self.labelframe21.destroy(),self.labelframe22.destroy(),self.labelframe23.destroy()])#self.labelframe20
        opciones6.add_command(label="Eliminar Items Compras", command=lambda:[self.widgets_modulos_eliminar_compras(),self.labelframe16.destroy(),self.labelframe17.destroy(),self.labelframe18.destroy(),self.labelframe19.destroy(),self.labelframe20.destroy(),self.labelframe21.destroy(),self.labelframe23.destroy()])#self.labelframe22
        opciones6.add_command(label="Eliminar Items Salidas", command=lambda:[self.widgets_modulos_eliminar_salidas(),self.labelframe16.destroy(),self.labelframe17.destroy(),self.labelframe18.destroy(),self.labelframe19.destroy(),self.labelframe20.destroy(),self.labelframe21.destroy(),self.labelframe22.destroy()])#self.labelframe23
        menubar1.add_cascade(label="Reportes", menu=opciones6)
        opciones5 = tk.Menu(menubar1)
        opciones5.add_command(label="Portada", command=lambda:[self.portada(),self.labelframe1.destroy(),self.labelframe15.destroy(),self.labelframe16.destroy(),self.labelframe17.destroy(),self.labelframe18.destroy(),self.labelframe19.destroy(),self.labelframe20.destroy(),self.labelframe21.destroy(),self.labelframe22.destroy(),self.labelframe23.destroy()])#self.labelframe18
        menubar1.add_cascade(label="Salir", menu=opciones5) 
        self.ventana1.mainloop()

    
    #0 # PORTADA--------------------------------------------------------------------------------------------------------#
    def portada(self):
        self.labelframe1=ttk.LabelFrame(self.ventana1, text="Datos Fiscales de la Empresa",borderwidth=20)
        self.labelframe1.grid(column=0,row=0,padx=5,ipadx=700,ipady=450)
        self.style = ttk.Style()
        self.style.theme_use("clam")
        self.style.configure("TNotebook.Tab", background="#FFFFFF", font=font.Font(family="verdana", size=10, weight="bold"))
        self.style =ttk.Style(self.labelframe1)
        self.style.configure('TLabelframe', background='#FFFFFF', weight="bold")
        self.label_portada=tkinter.Label(self.ventana1, text="G E S T I O N   D E   I N V E N T A R I O S",fg='Black',background='#FFFFFF',font=font.Font(family="Ebrima", size=15, weight = "bold"))
        self.label_portada.place(x=-25,y=25,width=500,height=90)
        self.portada1=tk.StringVar()
        self.portada01=tk.Entry(self.ventana1, 
                                textvariable=self.portada1, 
                                justify=tk.LEFT,
                                background="#FFFFFF",#00BFBF
                                foreground="black",
                                disabledbackground="#4d4d4d",
                                disabledforeground="#ffffff",
                                font=font.Font(family="Ebrima", size=10, weight = "bold"))
        self.portada01.place(x=20, 
                             y=100,
                             width=300,
                             height=25)
        self.portada2=tk.StringVar()
        self.portada02=tk.Entry(self.ventana1, 
                                 textvariable=self.portada2, 
                                 justify=tk.LEFT,
                                 background="#FFFFFF",#00BFBF
                                 foreground="black",
                                 disabledbackground="#4d4d4d",
                                 disabledforeground="#ffffff",
                                 font=font.Font(family="Ebrima", size=10, weight = "bold"))
        self.portada02.place(x=20, 
                             y=125,
                             width=450,
                             height=25)
        self.portada3=tk.StringVar()
        self.portada03=tk.Entry(self.ventana1, 
                                 textvariable=self.portada3,
                                 justify=tk.LEFT,
                                 background="#FFFFFF",#00BFBF
                                 foreground="black",
                                 disabledbackground="#4d4d4d",
                                 disabledforeground="#ffffff",
                                 font=font.Font(family="Ebrima", size=10, weight = "bold"))
        self.portada03.place(x=20, 
                             y=150,
                             width=850,
                             height=25)
        datos = self.base_datos_inventario.mostrar_datos_de_la_empresa()
        i = -1
        for dato in datos:
            i = i + 1
            self.portada1.set(datos[i][1])
            self.portada2.set(datos[i][2])
            self.portada3.set(datos[i][3])


    #0 # DATOS DE LA EMPRESA--------------------------------------------------------------------------------------------------------#
    def datos_de_la_empresa(self):
        self.labelframe15=ttk.LabelFrame(self.ventana1, text="Datos Fiscales de la Empresa",borderwidth=20)
        self.labelframe15.grid(column=0, row=0, padx=5,ipadx=700,ipady=450)
        self.style = ttk.Style()
        self.style.configure("MyEntry.TEntry",
                        # Fondo rojo.
                        fieldbackground="#FFFFFF",
                        # Color de texto azul.
                        foreground="#0000ff")
        self.style = ttk.Style()
        self.style.theme_use("clam")
        self.style.configure("TNotebook.Tab", background="#FFFFFF", font=font.Font(family="verdana", size=10, weight="bold"))
        self.style =ttk.Style(self.labelframe15)
        self.style.configure('TLabelframe', background='#FFFFFF', weight="bold")
        self.Razon_social=tk.StringVar()
        self.Razon_social1=ttk.Entry(self.labelframe15, textvariable=self.Razon_social,justify=tk.CENTER,style="MyEntry.TEntry")
        self.Razon_social1.place(x=100,y=140,width=300,height=40)
        self.label_Razon_social12=ttk.Label(self.labelframe15, text="Razon Social",foreground='blue',background='#FFFFFF',style="MyLabel.TLabel",font=font.Font(family="Corbel", size=10))
        self.label_Razon_social12.place(x=120,y=130)
        self.label_Razon_social156=ttk.Label(self.labelframe15, text="Ejem Empresa ABCDEFG, C.A.",foreground='blue',background='#FFFFFF',style="MyLabel.TLabel",font=font.Font(family="Corbel", size=10))
        self.label_Razon_social156.place(x=100,y=180)
        self.rif=tk.StringVar()
        self.rif2=ttk.Entry(self.labelframe15, textvariable=self.rif,justify=tk.CENTER,style="MyEntry.TEntry")
        self.rif2.place(x=100,y=230,width=300,height=40)
        self.label_rif=ttk.Label(self.labelframe15, text="R.I.F.",foreground='blue',background='#FFFFFF',style="MyLabel.TLabel",font=font.Font(family="Corbel", size=12))
        self.label_rif.place(x=120,y=220)
        self.label_rif45=ttk.Label(self.labelframe15, text="Ejem. J-123456789-0",foreground='blue',background='#FFFFFF',style="MyLabel.TLabel",font=font.Font(family="Corbel", size=8))
        self.label_rif45.place(x=100,y=270)
        self.direccion_fiscal=tk.StringVar()
        self.direccion_fiscal2=ttk.Entry(self.labelframe15, textvariable=self.direccion_fiscal,justify=tk.CENTER,style="MyEntry.TEntry")
        self.direccion_fiscal2.place(x=100,y=320,width=400,height=40)
        self.label_direccion_fiscal=ttk.Label(self.labelframe15, text="Direcc. Fiscal",foreground='blue',background='#FFFFFF',style="MyLabel.TLabel",font=font.Font(family="Corbel", size=12))
        self.label_direccion_fiscal.place(x=120,y=310)
        self.label_Razon_social114=ttk.Label(self.labelframe15, text="Ejme. Av. Caceres, Galpon 45, Zon Industrial Oviedo,Spain",foreground='blue',background='#FFFFFF',style="MyLabel.TLabel",font=font.Font(family="Corbel", size=8))
        self.label_Razon_social114.place(x=100,y=360)
        self.imgA = PhotoImage(file="E:/Contabilidad MySQL/Iconos/numero-uno.png")
        self.label_rifK=ttk.Label(self.labelframe15, image=self.imgA,compound="left",font=font.Font(family="Verdana", size=4))
        self.label_rifK.place(x=50,y=140)
        self.imgB = PhotoImage(file="E:/Contabilidad MySQL/Iconos/numero-2.png")
        self.label_rifK=ttk.Label(self.labelframe15, image=self.imgB,compound="left",font=font.Font(family="Verdana", size=4))
        self.label_rifK.place(x=50,y=230)
        self.imgC = PhotoImage(file="E:/Contabilidad MySQL/Iconos/numero-3.png")
        self.label_rifK=ttk.Label(self.labelframe15, image=self.imgC,compound="left",font=font.Font(family="Verdana", size=4))
        self.label_rifK.place(x=50,y=320)
        self.imgb2 = PhotoImage(file="E:/Contabilidad MySQL/Iconos/salvado.png")
        self.boton158 = tkinter.Button(self.labelframe15, image=self.imgb2, text="S A V E", fg='Black', bg='#FFFAF0',compound="left",font=font.Font(family="Verdana", size=8, weight = "bold"),command=lambda:[self.agregar_datos_empresa()])#self.info(),self.agregar_datos_empresa()
        self.boton158.place(x=50,y=-19,width=195,height=60)#690
        self.imgb3 = PhotoImage(file="E:/Contabilidad MySQL/Iconos/borrar.png")
        self.boton159 = tkinter.Button(self.labelframe15, image=self.imgb3,text="Borrar", fg='Black', bg='#FFFAF0',compound="left",font=font.Font(family="Verdana", size=8, weight = "bold"),command=lambda:[self.borrar_datos_empresas(),self.clear_pandas_datos_empresa()])
        self.boton159.place(x=240,y=-19,width=195,height=60)#880
        self.imgb4 = PhotoImage(file="E:/Contabilidad MySQL/Iconos/consulta.png")
        self.boton160 = tkinter.Button(self.labelframe15, image=self.imgb4, text="Consulta", fg='Black', bg='#FFFAF0',compound="left",font=font.Font(family="Verdana", size=8, weight = "bold"),command=lambda:[self.pandas_datos_empresa()])
        self.boton160.place(x=430,y=-19,width=195,height=60)
        self.imgb8 = PhotoImage(file="E:/Contabilidad MySQL/Iconos/eliminar_archivos.png")
        self.boton160 = tkinter.Button(self.labelframe15, image=self.imgb8, text="Eliminar", fg='Black', bg='#FFFAF0',compound="left",font=font.Font(family="Verdana", size=8, weight = "bold"),command=lambda:[self.Display_datos_de_la_empresa01()])
        self.boton160.place(x=620,y=-19,width=195,height=60)


    def Display_datos_de_la_empresa01(self):
        self.labelframeA1=ttk.LabelFrame(self.ventana1, text="Datos Fiscales de la Empresa",borderwidth=20)
        self.labelframeA1.grid(column=0, row=0, padx=5,ipadx=700,ipady=450)
        x = tk.messagebox.askquestion(message="¿Desea Consultar los Datos de la Empresa?", title="Información Fiscal")
        if x == 'yes':
            self.style = ttk.Style()
            self.style.theme_use("clam")
            self.style.configure("TNotebook.Tab", background="#FFFFFF", font=font.Font(family="verdana", size=10, weight="bold"))
            self.style =ttk.Style(self.labelframeA1)
            self.style.configure('TLabelframe', background='#FFFFFF', weight="bold")
            self.frm650 = ttk.LabelFrame(self.labelframeA1)#self.labelframeA1
            self.frm650.place(x=5,y=170,height=130)
            self.tvp=ttk.Treeview(self.frm650, columns =(1,2,3), show='headings')
            self.tvp.column(1, width=250,anchor='center')
            self.tvp.column(2, width=100,anchor='center')
            self.tvp.column(3, width=350,anchor='center')
            self.tvp.heading(1, text ='Razón Social')
            self.tvp.heading(2, text ='R.I.F')
            self.tvp.heading(3, text ='Direccion Fiscal')
            #self.tvp.heading(3, text ='Domicilio Fiscal')
            self.tvp.bind("<Double-1>", self.elimina_datos_empresa)
            self.tvp.bind()
            self.style =ttk.Style(self.tvp)
            self.style.configure('Treeview',  
                              background="#F0E68C",
                              foreground="black",
                              rowheight=15,
                              fieldbackground="white")
            self.style.map("Treeview",background=[('selected', 'green')])
            self.style.configure('Treeview', rowheight=13)
            self.tvp.pack(padx=10, pady=10, fill='both', expand=True)
            self.scrollbarp = ttk.Scrollbar(self.labelframeA1, orient=tk.VERTICAL, command=self.tvp.yview)
            self.tvp.configure(yscroll=self.scrollbarp.set)
            self.scrollbarp.place(x=710, y=201, height=87)
            self.boton_datos1 = tkinter.Button(self.labelframeA1, text="Consultar", fg='white', bg='#FF0000',compound="left",font=font.Font(family="Verdana", size=8, weight = "bold"),command=lambda:[self.Display_plan_de_cuentas_mysql300()])#text="🔍"
            self.boton_datos1.place(x=18,y=189,width=70, height=15)
            self.boton_datos2 = tkinter.Button(self.labelframeA1, text="Salir", fg='white', bg='#008B8B',compound="center",font=font.Font(family="Verdana", size=8, weight = "bold"),command=lambda:[self.exit1()])#self.INFO(),self.exit1()
            self.boton_datos2.place(x=87,y=189,width=50, height=15)

    def pandas_datos_empresa(self):
        datos = self.base_datoslc.mostrar_datos_de_la_empresa()                           
        x = tk.messagebox.askquestion(message="¿Desea Consultar Los Datos de la Empresa?", title="Comprobante de Retencion")
        if x == 'yes':
            df1p = pd.DataFrame(datos, columns=['id','Razon_social','rif ','direccion_fiscal','titulo1','titulo2','titulo3','titulo4'])
            filtered_df1p = df1p.drop(["id",'titulo1','titulo2','titulo3','titulo4'], axis=1)
            dfONEp = tabulate(filtered_df1p, headers="keys", tablefmt="fancy_grid")#rst,psql,pretty,fancy_grid,github,simple,plain
            self.table14 = tk.Text(self.labelframe15)
            self.table14.insert(tk.INSERT,dfONEp)
            self.table14.place(x=150,y=400,width=910,height=120)#(x=150, y=50,width=1500, height=700)

    def clear_pandas_datos_empresa(self):
        datos = self.base_datoslc.mostrar_datos_de_la_empresa()                           
        x = tk.messagebox.askquestion(message="¿Desea Borrar la Información?", title="Comprobante de Retencion")
        if x == 'yes':
            df1p = pd.DataFrame(datos, columns=['id','Razon_social','rif ','direccion_fiscal','titulo1','titulo2','titulo3','titulo4'])
            filtered_df1p = df1p.drop(["id",'titulo1','titulo2','titulo3','titulo4'], axis=1)
            dfONEp = tabulate(filtered_df1p, headers="keys", tablefmt="fancy_grid")#rst,psql,pretty,fancy_grid,github,simple,plain
            #self.table14 = tk.Text(self.labelframe15)
            self.table14.place(x=150,y=50,width=1100,height=200)
            self.table14.place_forget()


    def exit1(self):
        x = tk.messagebox.askquestion(message="¿Desea Salir?", title="Finalizar")
        if x == 'yes':
            self.labelframeA1.destroy()
    

    def borrar_datos_empresas(self):
        self.Razon_social1.delete(0, 'end')
        self.rif2.delete(0, 'end')
        self.direccion_fiscal2.delete(0, 'end')

    def Display_plan_de_cuentas_mysql300(self):
        datos = self.base_datoslc.mostrar_datos_de_la_empresa()
        self.tvp.delete(*self.tvp.get_children())
        i = -1
        x = tk.messagebox.askquestion(message="¿Consultar?", title="Consultar")
        if x == 'yes':
            for dato in datos:
                i = i + 1
                self.tvp.insert('',i, text=datos[i][1:2][0], values =datos[i][1:4])

    def elimina_datos_empresa(self,event):
        item = self.tvp.selection()[0]
        self.data1 = self.tvp.item(item)
        x = tk.messagebox.askquestion(message="¿Desea Eliminar los Datos de la Empresa?", title="Eliminar Asiento Contable")
        if x == 'yes':
            self.tvp.delete(item)
            self.base_datoslc.elimina_datos_empresa(self.data1['text'])

    def Quit_p(self):
        self.frm650.place_forget()
        self.tvp.pack_forget()
        self.scrollbarp.place_forget()
        self.boton_datos1.place_forget()
        self.boton_datos2.place_forget()

    def agregar_datos_empresa(self):
        Razon_social = self.Razon_social.get()
        rif = self.rif.get()
        direccion_fiscal = self.direccion_fiscal.get()
        titulo1 = 'Nº de Comprobante'
        titulo2 = 'Balance de Comprobación'
        titulo3 = 'Balance General'
        titulo4 = 'Estado de Resultado'
        x = tk.messagebox.askquestion(message="¿Desea Agregar los Datos de la Empresa?", title="Información Fiscal")
        if x == 'yes':
            datos = (Razon_social,rif,direccion_fiscal,titulo1,titulo2,titulo3,titulo4)
            if Razon_social and rif and direccion_fiscal and titulo1 and titulo2 and titulo3 and titulo4  !='':        
                self.base_datoslc.inserta_datos_empresa(Razon_social,rif,direccion_fiscal,titulo1,titulo2,titulo3,titulo4)

    def agregar_datos_empresa1(self):
        NComprobante1 ='0'
        RazonSocialAgenteRetencion = self.Razon_social.get()
        RIFAgenteRetencion = self.rif.get()
        NComprobante='0'
        x = tk.messagebox.askquestion(message="¿Desea Agregar los Datos de la Empresa?", title="Información Fiscal")
        if x == 'yes':
            datos = (NComprobante1,RazonSocialAgenteRetencion,RIFAgenteRetencion,NComprobante)
            if NComprobante1 and RazonSocialAgenteRetencion and RIFAgenteRetencion and NComprobante !='':        
                self.base_datoslc.inserta_datos_empresa(NComprobante1,RazonSocialAgenteRetencion,RIFAgenteRetencion,NComprobante)



    #------------------------------------------------------MODULOS DE INVENTARIOS----------------------------------------------------------------#
    def ingresar_codigo(self):
        self.labelframe17=ttk.LabelFrame(self.ventana1, text="Datos Fiscales de la Empresa",borderwidth=20)
        self.labelframe17.grid(column=0, row=0, padx=5, ipadx=716,ipady=450)
        self.style = ttk.Style()
        self.style.configure("MyEntry.TEntry",
                        # Fondo rojo.
                        fieldbackground="#FFFFFF",
                        # Color de texto azul.
                        foreground="#0000ff")
        self.style = ttk.Style()
        self.style.theme_use("clam")
        self.style.configure("TNotebook.Tab", background="#FFFFFF", font=font.Font(family="verdana", size=10, weight="bold"))
        self.style =ttk.Style(self.labelframe17)
        self.style.configure('TLabelframe', background='#FFFFFF', weight="bold")
        self.style = ttk.Style()
        self.style.configure("MyEntry.TEntry",
        # Blanco.
        fieldbackground="#FFFFFF",
        # Color de texto azul.
        foreground="#0000ff")
        #-----------------------------------------------------------------------------------------------------------------------#
        self.codigo_data_base=tk.StringVar()
        self.codigo_data_base1=ttk.Entry(self.labelframe17, textvariable=self.codigo_data_base,justify=tk.CENTER,style="MyEntry.TEntry")
        self.codigo_data_base1.place(x=0,y=80,width=200,height=30)#50
        self.label_Razon_social10=ttk.Label(self.labelframe17, text="Codigo:",foreground='blue',background='#FFFFFF',style="MyLabel.TLabel",font=font.Font(family="Corbel", size=10))
        self.label_Razon_social10.place(x=20,y=70)
        self.producto_data_base=tk.StringVar()
        self.producto_data_base1=ttk.Entry(self.labelframe17, textvariable=self.producto_data_base,justify=tk.CENTER)
        self.producto_data_base1.place(x=250,y=80,width=200,height=30)
        self.label_rif=ttk.Label(self.labelframe17, text="Producto:",foreground='blue',background='#FFFFFF',style="MyLabel.TLabel",font=font.Font(family="Corbel", size=10))
        self.label_rif.place(x=260,y=70)
        self.imgb2 = PhotoImage(file="E:/Contabilidad MySQL/Iconos/salvado.png")
        self.boton158 = tkinter.Button(self.labelframe17, image=self.imgb2, text="S A V E", fg='Black', bg='#FFFAF0',compound="left",font=font.Font(family="Verdana", size=8, weight = "bold"),command=lambda:[self.agregar_datos_codificacion_inentario()])
        self.boton158.place(x=-19,y=-19,width=195,height=60)#690.place(x=-19,y=-19,width=195,height=60)
        self.imgb3 = PhotoImage(file="E:/Contabilidad MySQL/Iconos/borrar.png")
        self.boton159 = tkinter.Button(self.labelframe17, image=self.imgb3,text="Borrar", fg='Black', bg='#FFFAF0',compound="left",font=font.Font(family="Verdana", size=8, weight = "bold"),command=lambda:[self.borrar_datos_codificacion()])
        self.boton159.place(x=171,y=-19,width=195,height=60)#880
        self.imgb4 = PhotoImage(file="E:/Contabilidad MySQL/Iconos/consulta.png")
        self.boton160 = tkinter.Button(self.labelframe17, image=self.imgb4, text="Listados de Codigos", fg='Black', bg='#FFFAF0',compound="left",font=font.Font(family="Verdana", size=8, weight = "bold"),command=lambda:[self.Display_datos_ingresar_codigos()])
        self.boton160.place(x=361,y=-19,width=195,height=60)


    def agregar_datos_codificacion_inentario(self):
        codigo = self.codigo_data_base.get()
        Producto = self.producto_data_base.get()
        x = tk.messagebox.askquestion(message="¿Desea Agregar Codificacion de Inventarios?", title="Codificacion Inventario")
        if x == 'yes':
            datos = (codigo,Producto)
            if codigo and Producto !='':        
                self.base_datos_inventario.inserta_codificacion_inventario(codigo,Producto)


    def borrar_datos_codificacion(self):
        self.codigo_data_base1.delete(0, 'end')
        self.producto_data_base1.delete(0, 'end')

    def Display_datos_ingresar_codigos(self):
        x = tk.messagebox.askquestion(message="¿Desea Consultar los Codificacion de los Productos?", title="Codificacion Inventario")
        if x == 'yes':
            self.style = ttk.Style()
            self.style.configure("MyEntry.TEntry",
                        # Fondo rojo.
                        fieldbackground="#FFFFFF",
                        # Color de texto azul.
                        foreground="#0000ff")
            self.style = ttk.Style()
            self.style.theme_use("clam")
            self.style.configure("TNotebook.Tab", background="#FFFFFF", font=font.Font(family="verdana", size=10, weight="bold"))
            self.style =ttk.Style(self.labelframe17)
            self.style.configure('TLabelframe', background='#FFFFFF', weight="bold")
            self.style = ttk.Style()
            self.style.configure("MyEntry.TEntry",
            # Blanco.
            fieldbackground="#FFFFFF",
            # Color de texto azul.
            foreground="#0000ff")
            self.frm850 = ttk.LabelFrame(self.labelframe17)
            self.frm850.place(x=5,y=170,height=330)#130
            self.tvp56=ttk.Treeview(self.frm850, columns =(1,2), show='headings')
            self.tvp56.column(1, width=100,anchor='center')
            self.tvp56.column(2, width=300,anchor='center')
            self.tvp56.heading(1, text ='Codificacion')
            self.tvp56.heading(2, text ='Producto')
            #self.tvp.heading(3, text ='Domicilio Fiscal')
            self.tvp56.bind("<Double-1>", self.elimina_codificacion2000)
            self.tvp56.bind()
            self.style =ttk.Style(self.tvp56)
            self.style.configure('Treeview',  
                              background="#F0E68C",
                              foreground="black",
                              rowheight=15,
                              fieldbackground="white")
            self.style.map("Treeview",background=[('selected', 'green')])
            self.style.configure('Treeview', rowheight=13)
            self.tvp56.pack(padx=10, pady=10, fill='both', expand=True)
            self.scrollbar56 = ttk.Scrollbar(self.labelframe17, orient=tk.VERTICAL, command=self.tvp56.yview)
            self.tvp56.configure(yscroll=self.scrollbar56.set)
            self.scrollbar56.place(x=420,y=201,height=287)
            self.boton_datos56 = tkinter.Button(self.labelframe17, text="Consultar", fg='white', bg='#FF0000',compound="left",font=font.Font(family="Verdana", size=8, weight = "bold"),command=lambda:[self.Display_codificacion_inventario()])#text="🔍"
            self.boton_datos56.place(x=18,y=189,width=100,height=15)
            #self.boton_po = tkinter.Button(self.labelframe1, text="🔍", fg='white', bg='#FF0000',compound="center",font=font.Font(family="Verdana", size=8, weight = "bold"),command=lambda:[self.Display_plan_de_cuentas_mysql300()])
            #self.boton_po.place(x=134,y=118,width=20,height=15)
            self.boton_datosjk = tkinter.Button(self.labelframe17, text="Salir", fg='white', bg='#008B8B',compound="center",font=font.Font(family="Verdana", size=8, weight = "bold"),command=lambda:[self.Quit_p()])
            self.boton_datosjk.place(x=115,y=189,width=100,height=15)

    def Display_codificacion_inventario(self):#Listo
        datos = self.base_datos_inventario.mostrar_codificacion()
        self.tvp56.delete(*self.tvp56.get_children())
        x = tk.messagebox.askquestion(message="¿Consultar?", title="Consultar")
        if x == 'yes':
            i = -1
            for dato in datos:
                i = i + 1
                self.tvp56.insert('',i, text=datos[i][1:2][0], values =datos[i][1:3])

    def Quit_p(self):#Listo
        self.frm850.place_forget()
        self.tvp56.pack_forget()
        self.scrollbar56.place_forget()
        self.boton_datos56.place_forget()
        self.boton_datosjk.place_forget()

    def elimina_codificacion2000(self,event):
        item = self.tvp56.selection()[0]
        self.data1 =  self.tvp56.item(item)
        x = tk.messagebox.askquestion(message="¿Desea Eliminar el Producto de la Base de Datos?", title="Eliminar Codificacion")
        if x == 'yes':
            self.tvp56.delete(item)
            self.base_datos_inventario.elimina_codificacion(self.data1['text'])


    #-----------------------------------------------------MODULO DE COMPRAS---------------------------------------------------#
    def widgets_modulos_compras(self):
        self.labelframe16=ttk.LabelFrame(self.ventana1, text="Modulo de Compras",borderwidth=20)
        self.labelframe16.grid(column=0, row=0, padx=5, ipadx=716,ipady=450)
        self.style = ttk.Style()
        self.style.configure("MyEntry.TEntry",
                        # Fondo rojo.
                        fieldbackground="#FFFFFF",
                        # Color de texto azul.
                        foreground="#0000ff")
        self.style = ttk.Style()
        self.style.theme_use("clam")
        self.style.configure("TNotebook.Tab", background="#FFFFFF", font=font.Font(family="verdana", size=10, weight="bold"))
        self.style =ttk.Style(self.labelframe16)
        self.style.configure('TLabelframe', background='#FFFFFF', weight="bold")
        self.style = ttk.Style()
        self.style.configure("MyEntry.TEntry",
        # Blanco.
        fieldbackground="#FFFFFF",
        # Color de texto azul.
        foreground="#0000ff")
        #-----------------------------------------------------------------------------------------------------------------------#
        self.labelpull=ttk.Label(self.labelframe16, text="REGISTROS ITEMS INVENTARIOS",background='#FFFFFF',font=font.Font(family="verdana", size=12, weight = "bold"))
        self.labelpull.place(x=800, y=50,height=20)
        self.label=ttk.Label(self.labelframe16, text="Registro del Inventario",anchor="center",foreground='white',background='#00BFBF',style="MyLabel.TLabel",font=font.Font(family="verdana", size=10))
        self.label.place(x=0,y=50,width=700,height=20)
        self.codigo_compras=tk.StringVar()
        self.entryc1=ttk.Entry(self.labelframe16, textvariable=self.codigo_compras,justify=tk.CENTER,style="MyEntry.TEntry")
        self.entryc1.place(x=0,y=80,width=200,height=30)#50
        self.producto_compras=tk.StringVar()
        self.entryc2=ttk.Entry(self.labelframe16, textvariable=self.producto_compras,justify=tk.CENTER,style="MyEntry.TEntry")
        self.entryc2.place(x=250,y=80,width=200,height=30)
        self.NFactura=tk.StringVar()
        self.entryc78=ttk.Entry(self.labelframe16, textvariable=self.NFactura,justify=tk.CENTER,style="MyEntry.TEntry")
        self.entryc78.place(x=500,y=80,width=200,height=30)
        self.Nlote=tk.StringVar()
        self.entryc79=ttk.Entry(self.labelframe16, textvariable=self.Nlote,justify=tk.CENTER,style="MyEntry.TEntry")
        self.entryc79.place(x=0,y=130,width=200,height=30)
        self.Fecha=tk.StringVar()
        self.entryc80=ttk.Entry(self.labelframe16, textvariable=self.Fecha,justify=tk.CENTER,style="MyEntry.TEntry")
        self.entryc80.place(x=250,y=130,width=200,height=30)
        self.Proveedor=tk.StringVar()
        self.entryc81=ttk.Entry(self.labelframe16, textvariable=self.Proveedor,justify=tk.CENTER,style="MyEntry.TEntry")
        self.entryc81.place(x=0,y=180,width=450,height=30)
        self.cantidades_compras=tk.StringVar()
        self.entryc3=ttk.Entry(self.labelframe16, textvariable=self.cantidades_compras,justify=tk.CENTER,style="MyEntry.TEntry")
        self.entryc3.place(x=0,y=230,width=200,height=30)
        self.costos_unitarios_compras=tk.StringVar()
        self.entryc4=ttk.Entry(self.labelframe16, textvariable=self.costos_unitarios_compras,justify=tk.CENTER,style="MyEntry.TEntry")
        self.entryc4.place(x=250,y=230,width=200,height=30)
        self.total_compras=tk.StringVar()
        self.entryc5=ttk.Entry(self.labelframe16, textvariable=self.total_compras,justify=tk.CENTER,style="MyEntry.TEntry",state=tk.DISABLED)
        self.entryc5.place(x=500,y=230,width=200,height=30)
        self.labelf1=ttk.Label(self.labelframe16, text="Codigo",foreground='blue',background='#FFFFFF',style="MyLabel.TLabel",font=font.Font(family="Corbel", size=10))
        self.labelf1.place(x=20,y=70)
        self.label=ttk.Label(self.labelframe16, text="Producto",foreground='blue',background='#FFFFFF',style="MyLabel.TLabel",font=font.Font(family="verdana", size=10))
        self.label.place(x=260,y=70)
        self.label=ttk.Label(self.labelframe16, text="Nº Factura",foreground='blue',background='#FFFFFF',style="MyLabel.TLabel",font=font.Font(family="verdana", size=10))
        self.label.place(x=520,y=70)
        self.label=ttk.Label(self.labelframe16, text="Nº Lote",foreground='blue',background='#FFFFFF',style="MyLabel.TLabel",font=font.Font(family="verdana", size=10))
        self.label.place(x=20,y=120)
        self.label=ttk.Label(self.labelframe16, text="Fecha",foreground='blue',background='#FFFFFF',style="MyLabel.TLabel",font=font.Font(family="verdana", size=10))
        self.label.place(x=260,y=120)
        self.label=ttk.Label(self.labelframe16, text="Proveedor",foreground='blue',background='#FFFFFF',style="MyLabel.TLabel",font=font.Font(family="verdana", size=10))
        self.label.place(x=20,y=170)
        self.label=ttk.Label(self.labelframe16, text="Cantidades",foreground='blue',background='#FFFFFF',style="MyLabel.TLabel",font=font.Font(family="verdana", size=10))
        self.label.place(x=20,y=220)
        self.label=ttk.Label(self.labelframe16, text="Costo Unitarios",foreground='blue',background='#FFFFFF',style="MyLabel.TLabel",font=font.Font(family="verdana", size=10))
        self.label.place(x=260,y=220)
        self.label=ttk.Label(self.labelframe16, text="Total",foreground='blue',background='#FFFFFF',style="MyLabel.TLabel",font=font.Font(family="verdana", size=10))
        self.label.place(x=520,y=220)
        self.imgb2 = PhotoImage(file="E:/Contabilidad MySQL/Iconos/salvado.png")
        self.boton158 = tkinter.Button(self.labelframe16, image=self.imgb2, text="S A V E", fg='Black', bg='#FFFAF0',compound="left",font=font.Font(family="Verdana", size=8, weight = "bold"),command=lambda:[self.modulo_compras()])
        self.boton158.place(x=50,y=-19,width=195,height=60)#690
        self.imgb3 = PhotoImage(file="E:/Contabilidad MySQL/Iconos/borrar.png")
        self.boton159 = tkinter.Button(self.labelframe16, image=self.imgb3,text="Borrar", fg='Black', bg='#FFFAF0',compound="left",font=font.Font(family="Verdana", size=8, weight = "bold"),command=lambda:[self.borrar_matriz_compras()])
        self.boton159.place(x=240,y=-19,width=195,height=60)#880
        self.imgb4 = PhotoImage(file="E:/Contabilidad MySQL/Iconos/consulta.png")
        self.boton160 = tkinter.Button(self.labelframe16, image=self.imgb4, text="Consulta Inventarios", fg='Black', bg='#FFFAF0',compound="left",font=font.Font(family="Verdana", size=8, weight = "bold"),command=lambda:[self.Display_base_datos()])
        self.boton160.place(x=430,y=-19,width=195,height=60)
        self.frm1 = ttk.LabelFrame(self.labelframe16)
        self.frm1.place(x=7, y=295,width=1225,height=275)
        self.tvcompras=ttk.Treeview(self.frm1, columns =(1,2,3,4,5,6,7,8,9), show='headings')
        self.tvcompras.column(1, width=90,anchor='center')
        self.tvcompras.column(2, width=90,anchor='center')
        self.tvcompras.column(3, width=90,anchor='center')
        self.tvcompras.column(4, width=170,anchor='center')
        self.tvcompras.column(5, width=90,anchor='center')
        self.tvcompras.column(6, width=100,anchor='center')
        self.tvcompras.column(7, width=100,anchor='center')
        self.tvcompras.column(8, width=60,anchor='center')
        self.tvcompras.column(9, width=100,anchor='center')
        self.tvcompras.heading(1, text ='NFactura')
        self.tvcompras.heading(2, text ='Nlote')
        self.tvcompras.heading(3, text ='Fecha')
        self.tvcompras.heading(4, text ='Proveedor')
        self.tvcompras.heading(5, text ='Codigo')
        self.tvcompras.heading(6, text ='Producto')
        self.tvcompras.heading(7, text ='Cantidades')
        self.tvcompras.heading(8, text ='C/U')
        self.tvcompras.heading(9, text ='Total')
        self.style =ttk.Style(self.tvcompras)
        self.style.configure('Treeview',  
                              background="#F0E68C",
                              foreground="black",
                              rowheight=15,
                              fieldbackground="white")
        self.style.map("Treeview",background=[('selected', 'green')])
        self.tvcompras.pack(padx=10, pady=10, fill='both', expand=True)
        self.scrollbar1 = ttk.Scrollbar(self.labelframe16, orient=tk.VERTICAL, command=self.tvcompras.yview)
        self.tvcompras.configure(yscroll=self.scrollbar1.set)
        self.scrollbar1.place(x=1220, y=313, height=259)
        self.hscrollbar = ttk.Scrollbar(self.labelframe16, orient=tk.HORIZONTAL)
        self.hscrollbar.config(command=self.tvcompras.xview)
        self.hscrollbar.place(x=6,y=570,width=1226,height=15)



    def modulo_compras(self):
        NFactura = self.NFactura.get()
        Nlote = self.Nlote.get()
        Fecha = self.Fecha.get()
        Proveedor = self.Proveedor.get()
        codigo=self.codigo_compras.get()
        Producto=self.producto_compras.get()
        Cantidades1 = float(self.cantidades_compras.get())
        Cantidades= '{:,.2f}'.format(Cantidades1).replace(",", "")
        CU1= float(self.costos_unitarios_compras.get())
        CU = ('{:,.2f}'.format(CU1).replace(",", ""))
        Total = float(Cantidades1) * float(CU1)
        Total1=self.total_compras.set(Total)
        x = tk.messagebox.askquestion(message="¿Desea procesar La Factura de Compras?", title="Procesar La Factura")
        if x == 'yes':
            self.tvcompras.insert('', 'end', values=(NFactura,
                                                     Nlote,
                                                     Fecha,
                                                     Proveedor,
                                                     codigo,
                                                    Producto,
                                                   '{:,.2f}'.format(Cantidades1).replace(",", "@").replace(".", ",").replace("@", "."),
                                                   '{:,.2f}'.format(CU1).replace(",", "@").replace(".", ",").replace("@", "."),
                                                   '{:,.2f}'.format(Total).replace(",", "@").replace(".", ",").replace("@", ".")))
            datos = (NFactura,Nlote,Fecha,Proveedor,codigo,Producto,Cantidades,CU)
            if NFactura and Nlote and Fecha and Proveedor and codigo and Producto and Cantidades and CU !='':        
                self.base_datos_inventario.insertar_compras(NFactura,Nlote,Fecha,Proveedor,codigo,Producto,Cantidades,CU)

    def borrar_matriz_compras(self):
        self.entryc1.delete(0,'end')
        self.entryc2.delete(0,'end')
        self.entryc3.delete(0,'end')
        self.entryc4.delete(0,'end')
        self.entryc5.delete(0,'end')
        self.entryc78.delete(0,'end')
        self.entryc79.delete(0,'end')
        self.entryc80.delete(0,'end')
        self.entryc81.delete(0,'end')
        #for item in self.tvcompras.get_children():
            #self.tvcompras.delete(item)


    def Display_base_datos(self):
        self.labelframeA7=ttk.LabelFrame(self.ventana1, text="Modulo de Compras",borderwidth=20)
        self.labelframeA7.grid(column=0, row=0, padx=5, ipadx=716,ipady=450)
        self.style = ttk.Style()
        self.style.configure("MyEntry.TEntry",
                        # Fondo rojo.
                        fieldbackground="#FFFFFF",
                        # Color de texto azul.
                        foreground="#0000ff")
        self.style = ttk.Style()
        self.style.theme_use("clam")
        self.style.configure("TNotebook.Tab", background="#FFFFFF", font=font.Font(family="verdana", size=10, weight="bold"))
        self.style =ttk.Style(self.labelframeA7)
        self.style.configure('TLabelframe', background='#FFFFFF', weight="bold")
        self.style = ttk.Style()
        self.style.configure("MyEntry.TEntry",
        # Blanco.
        fieldbackground="#FFFFFF",
        # Color de texto azul.
        foreground="#0000ff")
        self.frm4 = ttk.LabelFrame(self.labelframeA7)
        self.frm4.place(x=265,y=10,width=370,height=500)
        self.boton = tkinter.Button(self.labelframeA7, text="Consultar", fg='white', bg='#FF0000',compound="center",font=font.Font(family="Verdana", size=8, weight = "bold"),command=lambda:[self.Display_codificacion25()])
        self.boton.place(x=278,y=29,width=100, height=15)
        self.boton = tkinter.Button(self.labelframeA7, text="Borrar", fg='white', bg='Blue',compound="center",font=font.Font(family="Verdana", size=8, weight = "bold"),command=lambda:[self.clear_datos_productos()])
        self.boton.place(x=378,y=29,width=100, height=15)
        self.boton = tkinter.Button(self.labelframeA7, text="Salir", fg='white', bg='Blue',compound="center",font=font.Font(family="Verdana", size=8, weight = "bold"),command=lambda:[self.exit_registro_inventarios()])
        self.boton.place(x=478,y=29,width=100, height=15)
        self.tvk=ttk.Treeview(self.frm4, columns =(1,2), show='headings')
        self.tvk.column(1, width=110,anchor='sw')
        self.tvk.column(2, width=130,anchor='sw')
        self.tvk.heading(1, text ='Codigo')
        self.tvk.heading(2, text ='Producto')
        #self.tvk.bind("<Double-1>", self.elimina_codificacion2000)
        self.tvk.bind('<<TreeviewSelect>>', self.obtener_codificacion45)
        self.style1 =ttk.Style(self.tvk)
        self.style1.configure('Treeview',  background="#F0E68C",foreground="black",rowheight=15,fieldbackground="white")
        self.style.map("Treeview",background=[('selected', 'green')])
        self.tvk.pack(padx=10, pady=10, fill='both', expand=True)
        self.scrollbar = ttk.Scrollbar(self.labelframeA7, orient=tk.VERTICAL, command=self.tvk.yview)
        self.tvk.configure(yscroll=self.scrollbar.set)
        self.scrollbar.place(x=620,y=41,height=86)


    def clear_datos_productos(self):
        for item in self.tvk.get_children():
            self.tvk.delete(item)

    def exit_registro_inventarios(self):
        x = tk.messagebox.askquestion(message="¿Desea Salir?", title="Finalizar")
        if x == 'yes':
            self.labelframeA7.destroy()


    def Display_codificacion25(self):
        datos = self.base_datos_inventario.mostrar_codificacion2()
        self.tvk.delete(*self.tvk.get_children())
        x = tk.messagebox.askquestion(message="¿Desea Consultar Codificacion?", title="Codificacion")
        if x == 'yes':
            i = -1
            for dato in datos:
                i = i + 1
                self.tvk.insert('',i, text=datos[i][1:2][0], values =datos[i][1:3])


    def obtener_codificacion45(self,event):
        item = self.tvk.focus()
        self.data1 = self.tvk.item(item)
        self.codigo_compras.set(self.data1['text'])
        self.producto_compras.set((self.data1['values'][1]))

    
    #-----------------------------------------------------ELIMINAR ITEM MODULO DE COMPRAS---------------------------------------------------#
    def widgets_modulos_eliminar_compras(self):            
        self.labelframe22=ttk.LabelFrame(self.ventana1, text="Eliminar Items",borderwidth=20)            
        self.labelframe22.grid(column=0, row=0, padx=5, ipadx=716,ipady=450) 
        self.style = ttk.Style()
        self.style.configure("MyEntry.TEntry",
                        # Fondo rojo.
                        fieldbackground="#FFFFFF",
                        # Color de texto azul.
                        foreground="#0000ff")
        self.style = ttk.Style()
        self.style.theme_use("clam")
        self.style.configure("TNotebook.Tab", background="#FFFFFF", font=font.Font(family="verdana", size=10, weight="bold"))
        self.style =ttk.Style(self.labelframe1)
        self.style.configure('TLabelframe', background='#FFFFFF', weight="bold")
        self.style = ttk.Style()
        self.style.configure("MyEntry.TEntry",
        # Blanco.
        fieldbackground="#FFFFFF",
        # Color de texto azul.
        foreground="#0000ff")        
        self.labelpull=ttk.Label(self.labelframe22, text="ELIMINAR PRODUCTO POR Nº DE FACTURAS",background='#FFFFFF',font=font.Font(family="verdana", size=12, weight = "bold"))          
        self.labelpull.place(x=700, y=50,height=20)
        self.N_factura_c=tk.StringVar()
        self.entrys18=ttk.Entry(self.labelframe22, textvariable=self.N_factura_c,justify=tk.CENTER,style="MyEntry.TEntry")
        self.entrys18.place(x=0,y=80,width=200,height=30)#50
        self.labelf1=ttk.Label(self.labelframe22, text="Nº Factura",foreground='blue',background='#FFFFFF',style="MyLabel.TLabel",font=font.Font(family="Corbel", size=10))
        self.labelf1.place(x=20,y=70)
        self.imgb2 = PhotoImage(file="E:/Contabilidad MySQL/Iconos/consulta.png")
        self.boton158 = tkinter.Button(self.labelframe22, image=self.imgb2, text="Consulta", fg='Black', bg='#FFFAF0',compound="left",font=font.Font(family="Verdana", size=8, weight = "bold"),command=lambda:[self.Display_modulo_de_compras()])#self.Display_modulo_de_compras()
        self.boton158.place(x=-19,y=-19,width=195,height=60)#690
        self.imgb3 = PhotoImage(file="E:/Contabilidad MySQL/Iconos/borrar.png")
        self.boton159 = tkinter.Button(self.labelframe22, image=self.imgb3,text="Borrar", fg='Black', bg='#FFFAF0',compound="left",font=font.Font(family="Verdana", size=8, weight = "bold"),command=lambda:[self.clear_inv_compras()]) 
        self.boton159.place(x=171,y=-19,width=195,height=60)#880
        self.imgb4 = PhotoImage(file="E:/Contabilidad MySQL/Iconos/Balances.png")
        self.boton160 = tkinter.Button(self.labelframe22, image=self.imgb4, text="Consulta Producto", fg='Black', bg='#FFFAF0',compound="left",font=font.Font(family="Verdana", size=8, weight = "bold"),command=lambda:[self.Display_modulo_de_compras1()])   
        self.boton160.place(x=361,y=-19,width=195,height=60)
        self.frm1 = ttk.LabelFrame(self.labelframe22)           
        self.frm1.place(x=7, y=170,width=1225,height=415)           
        self.tveliminarcompras=ttk.Treeview(self.frm1, columns =(1,2,3,4,5,6,7,8), show='headings')                  
        self.tveliminarcompras.column(1, width=40,anchor='center')          
        self.tveliminarcompras.column(2, width=40,anchor='center')          
        self.tveliminarcompras.column(3, width=60,anchor='center')         
        self.tveliminarcompras.column(4, width=100,anchor='center')          
        self.tveliminarcompras.column(5, width=60,anchor='center')         
        self.tveliminarcompras.column(6, width=150,anchor='center')         
        self.tveliminarcompras.column(7, width=60,anchor='center')          
        self.tveliminarcompras.column(8, width=40,anchor='center') 
        self.tveliminarcompras.heading(1, text ='Nº Factura')                 
        self.tveliminarcompras.heading(2, text ='Nlote')            
        self.tveliminarcompras.heading(3, text ='Fecha')            
        self.tveliminarcompras.heading(4, text ='Proveedor')            
        self.tveliminarcompras.heading(5, text ='Codigo')           
        self.tveliminarcompras.heading(6, text ='Producto')         
        self.tveliminarcompras.heading(7, text ='Cantidades')           
        self.tveliminarcompras.heading(8, text ='C/U')          
        self.tveliminarcompras.bind("<Double-1>", self.elimina_item_compras)           
        self.style =ttk.Style(self.tveliminarcompras)           
        self.style.configure('Treeview',            
                              background="#F0E68C",         
                              foreground="black",           
                              rowheight=15,         
                              fieldbackground="white")          
        self.style.map("Treeview",background=[('selected', 'green')])           
        self.tveliminarcompras.pack(padx=10, pady=10, fill='both', expand=True)         
        self.scrollbar1 = ttk.Scrollbar(self.labelframe22, orient=tk.VERTICAL, command=self.tveliminarcompras.yview)            
        self.tveliminarcompras.configure(yscroll=self.scrollbar1.set)           
        self.scrollbar1.place(x=1220, y=202, height=370)            
        self.hscrollbar = ttk.Scrollbar(self.labelframe22, orient=tk.HORIZONTAL)            
        self.hscrollbar.config(command=self.tveliminarcompras.xview)            
        self.hscrollbar.place(x=6,y=570,width=1226,height=15)


    def Display_modulo_de_compras(self):#Listo
        NFactura = self.N_factura_c.get()
        NFactura = str("'" +  NFactura + "'")
        datos = self.base_datos_inventario.mostras_compras(NFactura)
        self.tveliminarcompras.delete(*self.tveliminarcompras.get_children())
        x = tk.messagebox.askquestion(message="¿Consultar Nº Factura?", title="Consultar")
        i = -1
        if x == 'yes':
            for dato in datos:
                i = i + 1
                self.tveliminarcompras.insert('',i, text=datos[i][1:2][0], values =datos[i][1:9])


    def elimina_item_compras(self,event):
        item = self.tveliminarcompras.selection()[0]
        self.data1 =  self.tveliminarcompras.item(item)
        x = tk.messagebox.askquestion(message="¿Desea Eliminar item?", title="Compras")
        if x == 'yes':
            self.tveliminarcompras.delete(item)
            self.base_datos_inventario.elimina_item_modulo_de_compra(self.data1['text'])

    def clear_inv_compras(self):
        self.entrys18.delete(0,'end')
        for item in self.tveliminarcompras.get_children():
            self.tveliminarcompras.delete(item)

    def Display_modulo_de_compras1(self):#Listo
        datos = self.base_datos_inventario.mostras_compras1()
        self.tveliminarcompras.delete(*self.tveliminarcompras.get_children())
        x = tk.messagebox.askquestion(message="¿Consultar?", title="Consultar")
        i = -1
        if x == 'yes':
            for dato in datos:
                i = i + 1
                self.tveliminarcompras.insert('',i, text=datos[i][1:2][0], values =datos[i][1:9])
          
    #--------------------------------------------------------MODULO DE SALIDAS-------------------------------------------------------#
    def widgets_modulos_salidas(self):
        self.labelframe18=ttk.LabelFrame(self.ventana1, text="Modulo de Compras",borderwidth=20)
        self.labelframe18.grid(column=0, row=0, padx=5, ipadx=716,ipady=450)
        self.style = ttk.Style()
        self.style.configure("MyEntry.TEntry",
                        # Fondo rojo.
                        fieldbackground="#FFFFFF",
                        # Color de texto azul.
                        foreground="#0000ff")
        self.style = ttk.Style()
        self.style.theme_use("clam")
        self.style.configure("TNotebook.Tab", background="#FFFFFF", font=font.Font(family="verdana", size=10, weight="bold"))
        self.style =ttk.Style(self.labelframe18)
        self.style.configure('TLabelframe', background='#FFFFFF', weight="bold")
        self.style = ttk.Style()
        self.style.configure("MyEntry.TEntry",
        # Blanco.
        fieldbackground="#FFFFFF",
        # Color de texto azul.
        foreground="#0000ff")
        #-----------------------------------------------------------------------------------------------------------------------#
        self.labelpull=ttk.Label(self.labelframe18, text="SALIDAS INVENTARIOS",background='#FFFFFF',font=font.Font(family="verdana", size=12, weight = "bold"))
        self.labelpull.place(x=700, y=50,height=20)
        self.label=ttk.Label(self.labelframe18, text="Registro del Inventario",anchor="center",foreground='white',background='#00BFBF',style="MyLabel.TLabel",font=font.Font(family="verdana", size=10))
        self.label.place(x=0,y=50,width=700,height=20)
        self.codigo_salidas=tk.StringVar()
        self.entrys1=ttk.Entry(self.labelframe18, textvariable=self.codigo_salidas,justify=tk.CENTER,style="MyEntry.TEntry")
        self.entrys1.place(x=0,y=80,width=200,height=30)#50
        self.producto_salidas=tk.StringVar()
        self.entrys2=ttk.Entry(self.labelframe18, textvariable=self.producto_salidas,justify=tk.CENTER,style="MyEntry.TEntry")
        self.entrys2.place(x=250,y=80,width=200,height=30)
        self.Valoración=tk.StringVar()
        self.entrys6=ttk.Entry(self.labelframe18, textvariable=self.Valoración,justify=tk.CENTER,style="MyEntry.TEntry")
        self.entrys6.place(x=500,y=80,width=200,height=30)
        self.NFactura_salidas=tk.StringVar()
        self.entrys123=ttk.Entry(self.labelframe18, textvariable=self.NFactura_salidas,justify=tk.CENTER,style="MyEntry.TEntry")
        self.entrys123.place(x=0,y=130,width=200,height=30)
        self.Nlote_salidas=tk.StringVar()
        self.entry124=ttk.Entry(self.labelframe18, textvariable=self.Nlote_salidas,justify=tk.CENTER,style="MyEntry.TEntry")
        self.entry124.place(x=250,y=130,width=200,height=30)
        self.cliente_salidas=tk.StringVar()
        self.entry126=ttk.Entry(self.labelframe18, textvariable=self.cliente_salidas,justify=tk.CENTER,style="MyEntry.TEntry")
        self.entry126.place(x=0,y=180,width=450,height=30)
        self.Fecha_salidas=tk.StringVar()
        self.entry788=ttk.Entry(self.labelframe18, textvariable=self.Fecha_salidas,justify=tk.CENTER,style="MyEntry.TEntry")
        self.entry788.place(x=0,y=230,width=200,height=30)
        self.cantidades_salidas=tk.StringVar()
        self.entrys3=ttk.Entry(self.labelframe18, textvariable=self.cantidades_salidas,justify=tk.CENTER,style="MyEntry.TEntry")
        self.entrys3.place(x=250,y=230,width=200,height=30)
        self.costos_unitarios_salidas=tk.StringVar()
        self.entrys4=ttk.Entry(self.labelframe18, textvariable=self.costos_unitarios_salidas,justify=tk.CENTER,style="MyEntry.TEntry")
        self.entrys4.place(x=500,y=230,width=200,height=30)
        self.total_salidas=tk.StringVar()
        self.entrys5=ttk.Entry(self.labelframe18, textvariable=self.total_salidas,justify=tk.CENTER,style="MyEntry.TEntry",state=tk.DISABLED)
        self.entrys5.place(x=0,y=280,width=200,height=30)
        self.labelf1=ttk.Label(self.labelframe18, text="Codigo",foreground='blue',background='#FFFFFF',style="MyLabel.TLabel",font=font.Font(family="Corbel", size=10))
        self.labelf1.place(x=20,y=70)
        self.label=ttk.Label(self.labelframe18, text="Producto",foreground='blue',background='#FFFFFF',style="MyLabel.TLabel",font=font.Font(family="verdana", size=10))
        self.label.place(x=260,y=70)
        self.label=ttk.Label(self.labelframe18, text="Valoracion",foreground='blue',background='#FFFFFF',style="MyLabel.TLabel",font=font.Font(family="verdana", size=10))
        self.label.place(x=520,y=70)
        self.label=ttk.Label(self.labelframe18, text="Nº Factura",foreground='blue',background='#FFFFFF',style="MyLabel.TLabel",font=font.Font(family="verdana", size=10))
        self.label.place(x=20,y=120)
        self.label=ttk.Label(self.labelframe18, text="Nº Lote",foreground='blue',background='#FFFFFF',style="MyLabel.TLabel",font=font.Font(family="verdana", size=10))
        self.label.place(x=260,y=120)
        self.label=ttk.Label(self.labelframe18, text="Cliente",foreground='blue',background='#FFFFFF',style="MyLabel.TLabel",font=font.Font(family="verdana", size=10))
        self.label.place(x=20,y=170)
        self.label=ttk.Label(self.labelframe18, text="Fecha",foreground='blue',background='#FFFFFF',style="MyLabel.TLabel",font=font.Font(family="verdana", size=10))
        self.label.place(x=20,y=220)
        self.label=ttk.Label(self.labelframe18, text="Cantidades",foreground='blue',background='#FFFFFF',style="MyLabel.TLabel",font=font.Font(family="verdana", size=10))
        self.label.place(x=260,y=220)
        self.label=ttk.Label(self.labelframe18, text="Costo Unitarios",foreground='blue',background='#FFFFFF',style="MyLabel.TLabel",font=font.Font(family="verdana", size=10))
        self.label.place(x=520,y=220)
        self.label=ttk.Label(self.labelframe18, text="Total",foreground='blue',background='#FFFFFF',style="MyLabel.TLabel",font=font.Font(family="verdana", size=10))
        self.label.place(x=0,y=270)
        self.imgb2 = PhotoImage(file="E:/Contabilidad MySQL/Iconos/salvado.png")
        self.boton158 = tkinter.Button(self.labelframe18, image=self.imgb2, text="S A V E", fg='Black', bg='#FFFAF0',compound="left",font=font.Font(family="Verdana", size=8, weight = "bold"),command=lambda:[self.modulo_salidas()])
        self.boton158.place(x=50,y=-19,width=195,height=60)#690
        self.imgb3 = PhotoImage(file="E:/Contabilidad MySQL/Iconos/borrar.png")
        self.boton159 = tkinter.Button(self.labelframe18, image=self.imgb3,text="Borrar", fg='Black', bg='#FFFAF0',compound="left",font=font.Font(family="Verdana", size=8, weight = "bold"),command=lambda:[self.borrar_matriz_salidas()])
        self.boton159.place(x=240,y=-19,width=195,height=60)#880
        self.imgb4654 = PhotoImage(file="E:/Contabilidad MySQL/Iconos/consulta.png")
        self.boton1608 = tkinter.Button(self.labelframe18, image=self.imgb4654, text="Consulta Inventarios", fg='Black', bg='#FFFAF0',compound="left",font=font.Font(family="Verdana", size=8, weight = "bold"),command=lambda:[self.Display_data_productos_salidas78()])
        self.boton1608.place(x=430,y=-19,width=195,height=60)
        self.imgb4 = PhotoImage(file="E:/Contabilidad MySQL/Iconos/calculadora.png")
        self.boton160 = tkinter.Button(self.labelframe18, image=self.imgb4, text="Valoracion de Inventrios", fg='Black', bg='#FFFAF0',compound="left",font=font.Font(family="Verdana", size=8, weight = "bold"),command=lambda:[self.valoracion1()])
        self.boton160.place(x=620,y=-19,width=195,height=60)
        self.frm1 = ttk.LabelFrame(self.labelframe18)
        self.frm1.place(x=7,y=320,width=1225,height=275)
        self.tvsalidas=ttk.Treeview(self.frm1, columns =(1,2,3,4,5,6,7,8,9), show='headings')
        self.tvsalidas.column(1, width=90,anchor='center')
        self.tvsalidas.column(2, width=80,anchor='center')
        self.tvsalidas.column(3, width=170,anchor='center')
        self.tvsalidas.column(4, width=60,anchor='center')
        self.tvsalidas.column(5, width=100,anchor='center')
        self.tvsalidas.column(6, width=100,anchor='center')
        self.tvsalidas.column(7, width=100,anchor='center')
        self.tvsalidas.column(8, width=100,anchor='center')
        self.tvsalidas.column(9, width=100,anchor='center')
        self.tvsalidas.heading(1, text ='NFactura')
        self.tvsalidas.heading(2, text ='Nlote')
        self.tvsalidas.heading(3, text ='Fecha')
        self.tvsalidas.heading(4, text ='cliente')
        self.tvsalidas.heading(5, text ='Codigo')
        self.tvsalidas.heading(6, text ='Producto')
        self.tvsalidas.heading(7, text ='Cantidades')
        self.tvsalidas.heading(8, text ='C/U')
        self.tvsalidas.heading(9, text ='Total')
        self.style =ttk.Style(self.tvsalidas)
        self.style.configure('Treeview',  
                              background="#F0E68C",
                              foreground="black",
                              rowheight=15,
                              fieldbackground="white")
        self.style.map("Treeview",background=[('selected', 'green')])
        self.tvsalidas.pack(padx=10, pady=10, fill='both', expand=True)
        self.scrollbar1 = ttk.Scrollbar(self.labelframe18, orient=tk.VERTICAL, command=self.tvsalidas.yview)
        self.tvsalidas.configure(yscroll=self.scrollbar1.set)
        self.scrollbar1.place(x=1220, y=338, height=257)
        self.hscrollbar = ttk.Scrollbar(self.labelframe18, orient=tk.HORIZONTAL)
        self.hscrollbar.config(command=self.tvsalidas.xview)
        self.hscrollbar.place(x=6,y=580,width=1226,height=15)


    def modulo_salidas(self):
        NFactura = self.NFactura_salidas.get()
        Nlote = self.Nlote_salidas.get()
        Fecha = self.Fecha_salidas.get()
        cliente = self.cliente_salidas.get()
        codigo=self.codigo_salidas.get()
        Producto=self.producto_salidas.get()
        Cantidades1 = float(self.cantidades_salidas.get())
        Cantidades= '{:,.2f}'.format(Cantidades1).replace(",", "")
        CU1= float(self.costos_unitarios_salidas.get())
        CU = '{:,.2}'.format(CU1).replace(",", "")
        Total = float(Cantidades1) * float(CU1)
        Total1=self.total_salidas.set(Total)
        x = tk.messagebox.askquestion(message="¿Desea procesar la Salida del Producto?", title="Salida de Inventario")
        if x == 'yes':
            self.tvsalidas.insert('', 'end', values=(NFactura,
                                                     Nlote,
                                                     Fecha,
                                                     cliente,
                                                     codigo,
                                                     Producto,
                                                    '{:,.2f}'.format(Cantidades1).replace(",", "@").replace(".", ",").replace("@", "."),
                                                    '{:,.2f}'.format(CU1).replace(",", "@").replace(".", ",").replace("@", "."),
                                                    '{:,.2f}'.format(Total).replace(",", "@").replace(".", ",").replace("@", ".")))
            datos = (NFactura,Nlote,Fecha,cliente,codigo,Producto,Cantidades,CU)
            if NFactura and Nlote and Fecha and cliente and codigo and Producto and Cantidades and CU !='':        
                self.base_datos_inventario.insertar_salidas(NFactura,Nlote,Fecha,cliente,codigo,Producto,Cantidades,CU)

    
    def exit_salidas_inventarios(self):
        x = tk.messagebox.askquestion(message="¿Desea Salir?", title="Finalizar")
        if x == 'yes':
            self.labelframeA8.destroy()


    def borrar_matriz_salidas(self):
        self.entrys123.delete(0,'end')
        self.entry124.delete(0,'end')
        self.entry788.delete(0,'end')
        self.entry126.delete(0,'end')
        self.entrys1.delete(0,'end')
        self.entrys2.delete(0,'end')
        self.entrys3.delete(0,'end')
        self.entrys4.delete(0,'end')
        self.entrys5.delete(0,'end')
        self.entrys6.delete(0,'end')
        #for item in self.tvcompras.get_children():
            #self.tvcompras.delete(item)


    def Display_data_productos_salidas78(self):
        self.labelframeA8=ttk.LabelFrame(self.ventana1, text="Almacen",borderwidth=20)
        self.labelframeA8.grid(column=0, row=0, padx=5, ipadx=716,ipady=320)
        self.style = ttk.Style()
        self.style.configure("MyEntry.TEntry",
                        # Fondo rojo.
                        fieldbackground="#FFFFFF",
                        # Color de texto azul.
                        foreground="#0000ff")
        self.style = ttk.Style()
        self.style.theme_use("clam")
        self.style.configure("TNotebook.Tab", background="#FFFFFF", font=font.Font(family="verdana", size=10, weight="bold"))
        self.style =ttk.Style(self.labelframeA8)
        self.style.configure('TLabelframe', background='#FFFFFF', weight="bold")
        self.style = ttk.Style()
        self.style.configure("MyEntry.TEntry",
        # Blanco.
        fieldbackground="#FFFFFF",
        # Color de texto azul.
        foreground="#0000ff")
        self.frm455 = ttk.LabelFrame(self.labelframeA8)
        self.frm455.place(x=265,y=10,width=370,height=130)
        self.tvAA=ttk.Treeview(self.frm455, columns =(1,2), show='headings')
        self.tvAA.column(1, width=110,anchor='sw')
        self.tvAA.column(2, width=130,anchor='sw')
        self.tvAA.heading(1, text ='Codigo')
        self.tvAA.heading(2, text ='Producto')
        self.tvAA.bind("<Double-1>", self.elimina_codificacion_salidas)
        self.tvAA.bind('<<TreeviewSelect>>', self.obtener_codificacion_salidas)
        self.style1 =ttk.Style(self.tvAA)
        self.style1.configure('Treeview',  background="#F0E68C",foreground="black",rowheight=15,fieldbackground="white")
        self.style.map("Treeview",background=[('selected', 'green')])
        self.tvAA.pack(padx=10, pady=10, fill='both', expand=True)
        self.scrollbar = ttk.Scrollbar(self.labelframeA8, orient=tk.VERTICAL, command=self.tvAA.yview)
        self.tvAA.configure(yscroll=self.scrollbar.set)
        self.scrollbar.place(x=620,y=41,height=86)
        self.boton = tkinter.Button(self.labelframeA8, text="Consultar", fg='white', bg='#FF0000',compound="center",font=font.Font(family="Verdana", size=8, weight = "bold"),command=lambda:[self.Display_codificacion_salidas()])
        self.boton.place(x=277,y=30,width=100, height=15)
        self.boton = tkinter.Button(self.labelframeA8, text="Borrar", fg='white', bg='Blue',compound="center",font=font.Font(family="Verdana", size=8, weight = "bold"),command=lambda:[self.clear_datos_productos1()])
        self.boton.place(x=378,y=30,width=100, height=15)
        self.boton = tkinter.Button(self.labelframeA8, text="Salidas", fg='white', bg='Green',compound="center",font=font.Font(family="Verdana", size=8, weight = "bold"),command=lambda:[self.exit_salidas_inventarios()])
        self.boton.place(x=478,y=30,width=100, height=15)



    def clear_datos_productos1(self):
        for item in self.tvAA.get_children():
            self.tvAA.delete(item)

    def Display_codificacion_salidas(self):
        datos = self.base_datos_inventario.mostrar_codificacion2()
        self.tvAA.delete(*self.tvAA.get_children())
        x = tk.messagebox.askquestion(message="¿Desea Consultar Codificacion?", title="Codificacion")
        if x == 'yes':
            i = -1
            for dato in datos:
                i = i + 1
                self.tvAA.insert('',i, text=datos[i][1:2][0], values =datos[i][1:3])


    def obtener_codificacion_salidas(self,event):
        item = self.tvAA.focus()
        self.data1 = self.tvAA.item(item)
        self.codigo_salidas.set(self.data1['text'])
        self.producto_salidas.set((self.data1['values'][1]))


    def elimina_codificacion_salidas(self,event):
        item = self.tvAA.selection()[0]
        self.data1 =  self.tvAA.item(item)
        x = tk.messagebox.askquestion(message="¿Desea Eliminar el Producto de la Base de Datos?", title="Eliminar Codificacion")
        if x == 'yes':
            self.tvAA.delete(item)
            self.base_datos_inventario.elimina_codificacion(self.data1['text'])

    
    def valoracion1(self):
        Valoración = (self.codigo_salidas.get(),)
        #Valoración = str("'" + Valoración + "'")
        Valoración = self.base_datos_inventario.valoracion(Valoración,Valoración,Valoración)  
        self.Valoración.set(Valoración)
    
    #----------------------------------------------------------ELIMINAR ITEM SALIDAS#----------------------------------------------------------#
    def widgets_modulos_eliminar_salidas(self):            
        self.labelframe23=ttk.LabelFrame(self.ventana1, text="Eliminar Items",borderwidth=20)            
        self.labelframe23.grid(column=0, row=0, padx=5, ipadx=716,ipady=450)
        self.style = ttk.Style()
        self.style.configure("MyEntry.TEntry",
                        # Fondo rojo.
                        fieldbackground="#FFFFFF",
                        # Color de texto azul.
                        foreground="#0000ff")
        self.style = ttk.Style()
        self.style.theme_use("clam")
        self.style.configure("TNotebook.Tab", background="#FFFFFF", font=font.Font(family="verdana", size=10, weight="bold"))
        self.style =ttk.Style(self.labelframe23)
        self.style.configure('TLabelframe', background='#FFFFFF', weight="bold")
        self.style = ttk.Style()
        self.style.configure("MyEntry.TEntry",
        # Blanco.
        fieldbackground="#FFFFFF",
        # Color de texto azul.
        foreground="#0000ff")   
        self.labelpull=ttk.Label(self.labelframe23, text="ELIMINAR PRODUCTO POR Nº FACTURA",background='#FFFFFF',font=font.Font(family="verdana", size=12, weight = "bold"))          
        self.labelpull.place(x=600, y=-17,height=20)
        self.N_Factura_s=tk.StringVar()
        self.entrys187=ttk.Entry(self.labelframe23, textvariable=self.N_Factura_s,justify=tk.CENTER,style="MyEntry.TEntry")
        self.entrys187.place(x=0,y=80,width=200,height=30)#50
        self.labelf1=ttk.Label(self.labelframe23, text="Nº Factura",foreground='blue',background='#FFFFFF',style="MyLabel.TLabel",font=font.Font(family="Corbel", size=10))
        self.labelf1.place(x=20,y=70)
        self.imgb2 = PhotoImage(file="E:/Contabilidad MySQL/Iconos/consulta.png")
        self.boton158 = tkinter.Button(self.labelframe23, image=self.imgb2, text="Consulta", fg='Black', bg='#FFFAF0',compound="left",font=font.Font(family="Verdana", size=8, weight = "bold"),command=lambda:[self.Display_modulo_de_salidas()])#self.libro_compra(),self.n_comprobante1()
        self.boton158.place(x=-19,y=-19,width=195,height=60)#690
        self.imgb3 = PhotoImage(file="E:/Contabilidad MySQL/Iconos/borrar.png")
        self.boton159 = tkinter.Button(self.labelframe23, image=self.imgb3,text="Borrar", fg='Black', bg='#FFFAF0',compound="left",font=font.Font(family="Verdana", size=8, weight = "bold"),command=lambda:[self.clear_inv_salidas()])
        self.boton159.place(x=171,y=-19,width=195,height=60)#880
        self.imgb4 = PhotoImage(file="E:/Contabilidad MySQL/Iconos/Balances.png")
        self.boton160 = tkinter.Button(self.labelframe23, image=self.imgb4, text="Consulta Productos", fg='Black', bg='#FFFAF0',compound="left",font=font.Font(family="Verdana", size=8, weight = "bold"),command=lambda:[self.Display_modulo_de_salidas1()])
        self.boton160.place(x=361,y=-19,width=195,height=60)
        self.frm1 = ttk.LabelFrame(self.labelframe23)           
        self.frm1.place(x=7, y=170,width=1225,height=415)           
        self.tveliminarsalidas=ttk.Treeview(self.frm1, columns =(1,2,3,4,5,6,7,8), show='headings')                  
        self.tveliminarsalidas.column(1, width=90,anchor='center')          
        self.tveliminarsalidas.column(2, width=90,anchor='center')          
        self.tveliminarsalidas.column(3, width=170,anchor='center')         
        self.tveliminarsalidas.column(4, width=90,anchor='center')          
        self.tveliminarsalidas.column(5, width=100,anchor='center')         
        self.tveliminarsalidas.column(6, width=100,anchor='center')         
        self.tveliminarsalidas.column(7, width=60,anchor='center')          
        self.tveliminarsalidas.column(8, width=100,anchor='center') 
        self.tveliminarsalidas.heading(1, text ='Nº Factura')                 
        self.tveliminarsalidas.heading(2, text ='Nlote')            
        self.tveliminarsalidas.heading(3, text ='Fecha')            
        self.tveliminarsalidas.heading(4, text ='Proveedor')            
        self.tveliminarsalidas.heading(5, text ='Codigo')           
        self.tveliminarsalidas.heading(6, text ='Producto')         
        self.tveliminarsalidas.heading(7, text ='Cantidades')           
        self.tveliminarsalidas.heading(8, text ='C/U')          
        self.tveliminarsalidas.bind("<Double-1>", self.elimina_item_salidas)           
        self.style =ttk.Style(self.tveliminarsalidas)           
        self.style.configure('Treeview',            
                              background="#F0E68C",         
                              foreground="black",           
                              rowheight=15,         
                              fieldbackground="white")          
        self.style.map("Treeview",background=[('selected', 'green')])           
        self.tveliminarsalidas.pack(padx=10, pady=10, fill='both', expand=True)         
        self.scrollbar1 = ttk.Scrollbar(self.labelframe23, orient=tk.VERTICAL, command=self.tveliminarsalidas.yview)            
        self.tveliminarsalidas.configure(yscroll=self.scrollbar1.set)           
        self.scrollbar1.place(x=1220, y=202, height=370)            
        self.hscrollbar = ttk.Scrollbar(self.labelframe23, orient=tk.HORIZONTAL)            
        self.hscrollbar.config(command=self.tveliminarsalidas.xview)            
        self.hscrollbar.place(x=6,y=570,width=1226,height=15)
    
    def Display_modulo_de_salidas(self):#Listo
        NFactura = self.N_Factura_s.get()
        NFactura = str("'" +  NFactura + "'")
        datos = self.base_datos_inventario.mostras_salidas(NFactura)
        self.tveliminarsalidas.delete(*self.tveliminarsalidas.get_children())
        x = tk.messagebox.askquestion(message="¿Consultar?", title="Consultar")
        i = -1
        if x == 'yes':
            for dato in datos:
                i = i + 1
                self.tveliminarsalidas.insert('',i, text=datos[i][1:2][0], values =datos[i][1:9])


    def elimina_item_salidas(self,event):
        item = self.tveliminarsalidas.selection()[0]
        self.data1 =  self.tveliminarsalidas.item(item)
        x = tk.messagebox.askquestion(message="¿Desea Eliminar item?", title="Compras")
        if x == 'yes':
            self.tveliminarsalidas.delete(item)
            self.base_datos_inventario.elimina_item_modulo_de_salida(self.data1['text'])

    def clear_inv_salidas(self):
        self.entrys187.delete(0,'end')
        for item in self.tveliminarsalidas.get_children():
            self.tveliminarsalidas.delete(item)


    def Display_modulo_de_salidas1(self):#Listo
        datos = self.base_datos_inventario.mostras_salidas1()
        self.tveliminarsalidas.delete(*self.tveliminarsalidas.get_children())
        x = tk.messagebox.askquestion(message="¿Consultar?", title="Consultar")
        i = -1
        if x == 'yes':
            for dato in datos:
                i = i + 1
                self.tveliminarsalidas.insert('',i, text=datos[i][1:2][0], values =datos[i][1:9])

    #--------------------------------------------------------REPORTE INVENTARIOS ALL ITEMS-------------------------------------------------------# 
    def widgets_reportes_all_inventarios(self):
        self.labelframe19=ttk.LabelFrame(self.ventana1, text="Reporte Inventario All Items",borderwidth=20)
        self.labelframe19.grid(column=0, row=0, padx=5, ipadx=716,ipady=450)
        self.style = ttk.Style()
        self.style.configure("MyEntry.TEntry",
                        # Fondo rojo.
                        fieldbackground="#FFFFFF",
                        # Color de texto azul.
                        foreground="#0000ff")
        self.style = ttk.Style()
        self.style.theme_use("clam")
        self.style.configure("TNotebook.Tab", background="#FFFFFF", font=font.Font(family="verdana", size=10, weight="bold"))
        self.style =ttk.Style(self.labelframe19)
        self.style.configure('TLabelframe', background='#FFFFFF', weight="bold")
        self.style = ttk.Style()
        self.style.configure("MyEntry.TEntry",
        # Blanco.
        fieldbackground="#FFFFFF",
        # Color de texto azul.
        foreground="#0000ff")
        self.labelpull=ttk.Label(self.labelframe19, text="INVENTARIO FINAL DE MERCANCIAS",background='#FFFFFF',font=font.Font(family="verdana", size=12, weight = "bold"))
        self.labelpull.place(x=600, y=50,height=20)
        self.imgb2 = PhotoImage(file="E:/Contabilidad MySQL/Iconos/consulta.png")
        self.boton158 = tkinter.Button(self.labelframe19, image=self.imgb2, text="Consulta", fg='Black', bg='#FFFAF0',compound="left",font=font.Font(family="Verdana", size=8, weight = "bold"),command=lambda:[self.generar_all_itmes_inventory()])
        self.boton158.place(x=-19,y=-19,width=195,height=60)#690
        self.imgb3 = PhotoImage(file="E:/Contabilidad MySQL/Iconos/borrar.png")
        self.boton159 = tkinter.Button(self.labelframe19, image=self.imgb3,text="Borrar", fg='Black', bg='#FFFAF0',compound="left",font=font.Font(family="Verdana", size=8, weight = "bold"),command=lambda:[self.clear_inventario_final()])
        self.boton159.place(x=171,y=-19,width=195,height=60)#880
        self.imgb4 = PhotoImage(file="E:/Contabilidad MySQL/Iconos/excel.png")
        self.boton160 = tkinter.Button(self.labelframe19, image=self.imgb4, text="Excel", fg='Black', bg='#FFFAF0',compound="left",font=font.Font(family="Verdana", size=8, weight = "bold"),command=lambda:[self.generar_all_itmes_inventory_excel('worksheet')])
        self.boton160.place(x=361,y=-19,width=195,height=60)
        self.frm1 = ttk.LabelFrame(self.labelframe19)
        self.frm1.place(x=7,y=145,width=1225,height=275)#place(x=7,y=295,width=1225,height=275)
        self.tvreportInv=ttk.Treeview(self.frm1, columns =(1,2,3,4,5,6,7,8,9,10,11,12,13,14), show='headings')
        self.tvreportInv.column(1, width=100,anchor='center')
        self.tvreportInv.column(2, width=100,anchor='center')
        self.tvreportInv.column(3, width=80,anchor='center')
        self.tvreportInv.column(4, width=60,anchor='center')
        self.tvreportInv.column(5, width=80,anchor='center')
        self.tvreportInv.column(6, width=80,anchor='center')
        self.tvreportInv.column(7, width=60,anchor='center')
        self.tvreportInv.column(8, width=80,anchor='center')
        self.tvreportInv.column(9, width=80,anchor='center')
        self.tvreportInv.column(10, width=60,anchor='center')
        self.tvreportInv.column(11, width=80,anchor='center')
        self.tvreportInv.column(12, width=80,anchor='center')
        self.tvreportInv.column(13, width=60,anchor='center')
        self.tvreportInv.column(14, width=80,anchor='center')
        self.tvreportInv.heading(1, text ='Codigo')
        self.tvreportInv.heading(2, text ='Producto')
        self.tvreportInv.heading(3, text ='Cantidades')
        self.tvreportInv.heading(4, text ='C/U')
        self.tvreportInv.heading(5, text ='Total')
        self.tvreportInv.heading(6, text ='Cantidades')
        self.tvreportInv.heading(7, text ='C/U')
        self.tvreportInv.heading(8, text ='Total')
        self.tvreportInv.heading(9, text ='Cantidades')
        self.tvreportInv.heading(10, text ='C/U')
        self.tvreportInv.heading(11, text ='Total')
        self.tvreportInv.heading(12, text ='Cantidades')
        self.tvreportInv.heading(13, text ='C/U')
        self.tvreportInv.heading(14, text ='Total')
        self.style =ttk.Style(self.tvreportInv)
        self.style.configure('Treeview',  
                              background="#F0E68C",
                              foreground="black",
                              rowheight=15,
                              fieldbackground="white")
        self.style.map("Treeview",background=[('selected', 'green')])
        self.tvreportInv.pack(padx=10, pady=10, fill='both', expand=True)
        self.scrollbar1 = ttk.Scrollbar(self.labelframe19, orient=tk.VERTICAL, command=self.tvreportInv.yview)
        self.tvreportInv.configure(yscroll=self.scrollbar1.set)
        self.scrollbar1.place(x=1220, y=163, height=259)
        self.hscrollbar = ttk.Scrollbar(self.labelframe19, orient=tk.HORIZONTAL)
        self.hscrollbar.config(command=self.tvreportInv.xview)
        self.hscrollbar.place(x=6,y=420,width=1226,height=15)


    def clear_inventario_final(self):
        for item in self.tvreportInv.get_children():
            self.tvreportInv.delete(item)



    def widgets_reportes_all_inventario1(self):
        self.style = ttk.Style()
        self.style.configure("MyEntry.TEntry",
                        # Fondo rojo.
                        fieldbackground="#FFFFFF",
                        # Color de texto azul.
                        foreground="#0000ff")
        self.style = ttk.Style()
        self.style.theme_use("clam")
        self.style.configure("TNotebook.Tab", background="#FFFFFF", font=font.Font(family="verdana", size=10, weight="bold"))
        self.style =ttk.Style(self.labelframe19)
        self.style.configure('TLabelframe', background='#FFFFFF', weight="bold")
        self.style = ttk.Style()
        self.style.configure("MyEntry.TEntry",
        # Blanco.
        fieldbackground="#FFFFFF",
        # Color de texto azul.
        foreground="#0000ff")
        self.label=ttk.Label(self.labelframe19, text="Inventario Inicial",foreground='blue',background='#ADD8E6',font=font.Font(family="verdana", size=15, weight = "bold"),anchor='center')
        self.label.place(x=238,y=148,width=245,height=30)
        self.label=ttk.Label(self.labelframe19, text="Compras",foreground='blue',background='#6495ED',font=font.Font(family="verdana", size=15, weight = "bold"),anchor='center')
        self.label.place(x=480,y=148,width=245,height=30)
        self.label=ttk.Label(self.labelframe19, text="Salidas",foreground='blue',background='#00BFFF',font=font.Font(family="verdana", size=15, weight = "bold"),anchor='center')
        self.label.place(x=725,y=148,width=245,height=30)
        self.label=ttk.Label(self.labelframe19, text="Inventario Final",foreground='blue',background='#1E90FF',font=font.Font(family="verdana", size=15, weight = "bold"),anchor='center')
        self.label.place(x=970,y=148,width=250,height=30)


    def generar_all_itmes_inventory(self):
        all_itmes = self.base_datos_inventario.reportes_all_inventario()  
        self.tvreportInv.delete(*self.tvreportInv.get_children())
        x = tk.messagebox.askquestion(message="¿Desea generar Reportes de Invenatrios?", title="Rpeortes de Inventrios")
        if x == 'yes':
            for i in all_itmes:
                codigo = (i[0])
                Producto = (i[1])
                Cantidades_iniciales = (i[2])       
                CU_iniciales = (i[3])       
                Total_iniciales = (i[4])       
                Cantidades = (i[5])
                CU = (i[6])
                Total1 = (i[7])
                Cantidades2 = (i[8])
                CU_Salidas = ((i[9]))
                Total2 = (i[10])
                Cantidades3 = (i[11])
                CU3 = (i[12])
                Total3 = (i[13])
                self.tvreportInv.insert('', 'end', values=(codigo,
                                                           Producto,
                                                           '{:,.2f}'.format(Cantidades_iniciales).replace(",", "@").replace(".", ",").replace("@", "."),
                                                            '{:,.2f}'.format(CU_iniciales).replace(",", "@").replace(".", ",").replace("@", "."),
                                                            '{:,.2f}'.format(Total_iniciales).replace(",", "@").replace(".", ",").replace("@", "."),
                                                            '{:,.2f}'.format(Cantidades).replace(",", "@").replace(".", ",").replace("@", "."),
                                                            '{:,.2f}'.format(CU).replace(",", "@").replace(".", ",").replace("@", "."),
                                                            '{:,.2f}'.format(Total1).replace(",", "@").replace(".", ",").replace("@", "."),
                                                            '{:,.2f}'.format(Cantidades2).replace(",", "@").replace(".", ",").replace("@", "."),
                                                            CU_Salidas,
                                                            '{:,.2f}'.format(Total2).replace(",", "@").replace(".", ",").replace("@", "."),
                                                            '{:,.2f}'.format(Cantidades3).replace(",", "@").replace(".", ",").replace("@", "."),
                                                            CU3,
                                                            '{:,.2f}'.format(Total3).replace(",", "@").replace(".", ",").replace("@", ".")))    
                

    def generar_all_itmes_inventory_excel(self,worksheet, side=None, blank=True):
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.merge_cells('B1:O1')
        self.ws['B1'] = 'I N V E N T A R I O S'
        self.ws['B1'].alignment = Alignment(horizontal='center', vertical='center')
        self.ws.merge_cells('B3')
        self.ws['B3'] = 'Codigo'
        self.ws['B3'].alignment = Alignment(horizontal='center', vertical='center')
        self.ws.merge_cells('C3')
        self.ws['C3'] =  'Producto'
        self.ws['C3'].alignment = Alignment(horizontal='center', vertical='center')
        self.ws.merge_cells('D2:F2')
        self.ws['D2'] = 'Inventario Inicial'
        self.ws['D2'].alignment = Alignment(horizontal='center', vertical='center')
        self.ws.merge_cells('G2:I2')
        self.ws['G2'] = 'Compras'
        self.ws['G2'].alignment = Alignment(horizontal='center', vertical='center')
        self.ws.merge_cells('J2:L2')
        self.ws['J2'] = 'Salidas'
        self.ws['J2'].alignment = Alignment(horizontal='center', vertical='center')
        self.ws.merge_cells('M2:O2')
        self.ws['M2'] = 'Inventario Final'
        self.ws['M2'].alignment = Alignment(horizontal='center', vertical='center')
        self.ws['D3'] = 'Cantidades'
        self.ws['D3'].alignment = Alignment(horizontal='center', vertical='center')
        self.ws['E3'] = 'C/U'
        self.ws['E3'].alignment = Alignment(horizontal='center', vertical='center')
        self.ws['F3'] = 'Total'
        self.ws['F3'].alignment = Alignment(horizontal='center', vertical='center')
        self.ws['G3'] = 'Cantidades'
        self.ws['G3'].alignment = Alignment(horizontal='center', vertical='center')
        self.ws['H3'] = 'C/U'
        self.ws['H3'].alignment = Alignment(horizontal='center', vertical='center')
        self.ws['I3'] = 'Total'
        self.ws['I3'].alignment = Alignment(horizontal='center', vertical='center')
        self.ws['J3'] = 'Cantidades'
        self.ws['J3'].alignment = Alignment(horizontal='center', vertical='center')
        self.ws['K3'] = 'C/U'
        self.ws['K3'].alignment = Alignment(horizontal='center', vertical='center')
        self.ws['L3'] = 'Total'
        self.ws['L3'].alignment = Alignment(horizontal='center', vertical='center')
        self.ws['M3'] = 'Cantidades'
        self.ws['M3'].alignment = Alignment(horizontal='center', vertical='center')
        self.ws['N3'] = 'C/U'
        self.ws['N3'].alignment = Alignment(horizontal='center', vertical='center')
        self.ws['O3'] = 'Total'
        self.ws['O3'].alignment = Alignment(horizontal='center', vertical='center')
        for i, item in enumerate(self.tvreportInv.get_children()):
            values = list(self.tvreportInv.item(item)['values'])
            self.ws.cell(row=i+4, column=1, value=i+1)
            for j, value in enumerate(values):
                self.ws.cell(row=i+4, column=j+2, value=value)
        messagebox.showinfo(title='Inventarios', message='Descargando Inventarios')
        side = Side(border_style='thin', color='FF0000')
        self.set_border(self.ws, side)
        self.wb.save('Inventarios_Item.xlsx')
        self.wb.close()
    

    #--------------------------------------------------------REPORTE INVENTARIOS ONE ITEMS-------------------------------------------------------# 
    def widgets_reportes_one_inventarios(self):
        self.labelframe20=ttk.LabelFrame(self.ventana1, text="Reporte Inventario All Items",borderwidth=20)
        self.labelframe20.grid(column=0, row=0, padx=5, ipadx=716,ipady=450)
        self.style = ttk.Style()
        self.style.configure("MyEntry.TEntry",
                        # Fondo rojo.
                        fieldbackground="#FFFFFF",
                        # Color de texto azul.
                        foreground="#0000ff")
        self.style = ttk.Style()
        self.style.theme_use("clam")
        self.style.configure("TNotebook.Tab", background="#FFFFFF", font=font.Font(family="verdana", size=10, weight="bold"))
        self.style =ttk.Style(self.labelframe20)
        self.style.configure('TLabelframe', background='#FFFFFF', weight="bold")
        self.style = ttk.Style()
        self.style.configure("MyEntry.TEntry",
        # Blanco.
        fieldbackground="#FFFFFF",
        # Color de texto azul.
        foreground="#0000ff")
        self.labelpull=ttk.Label(self.labelframe20, text="REPORTES DE INVENTARIOS POR PRODUCTO",background='#FFFFFF',font=font.Font(family="verdana", size=12, weight = "bold"))
        self.labelpull.place(x=700,y=50,height=20)
        self.codigo_report=tk.StringVar()
        self.entrycp=ttk.Entry(self.labelframe20, textvariable=self.codigo_report,justify=tk.CENTER,style="MyEntry.TEntry")
        self.entrycp.place(x=0,y=80,width=200,height=30)#50
        self.labelf1=ttk.Label(self.labelframe20, text="Codigo",foreground='blue',background='#FFFFFF',style="MyLabel.TLabel",font=font.Font(family="Corbel", size=10))
        self.labelf1.place(x=20,y=70)
        self.imgb2 = PhotoImage(file="E:/Contabilidad MySQL/Iconos/consulta.png")
        self.boton158 = tkinter.Button(self.labelframe20, image=self.imgb2, text="Consulta", fg='Black', bg='#FFFAF0',compound="left",font=font.Font(family="Verdana", size=8, weight = "bold"),command=lambda:[self.generar_one_itmes_inventory()])
        self.boton158.place(x=-19,y=-19,width=195,height=60)#690
        self.imgb3 = PhotoImage(file="E:/Contabilidad MySQL/Iconos/borrar.png")
        self.boton159 = tkinter.Button(self.labelframe20, image=self.imgb3,text="Borrar", fg='Black', bg='#FFFAF0',compound="left",font=font.Font(family="Verdana", size=8, weight = "bold"),command=lambda:[self.clear_datos_one_productos1()])
        self.boton159.place(x=171,y=-19,width=195,height=60)#880
        self.imgb4 = PhotoImage(file="E:/Contabilidad MySQL/Iconos/excel.png")
        self.boton160 = tkinter.Button(self.labelframe20, image=self.imgb4, text="Excel", fg='Black', bg='#FFFAF0',compound="left",font=font.Font(family="Verdana", size=8, weight = "bold"),command=lambda:[self.generar_one_itmes_inventory_excel('worksheet')])
        self.boton160.place(x=361,y=-19,width=195,height=60)
        self.frm1 = ttk.LabelFrame(self.labelframe20)
        self.frm1.place(x=7,y=145,width=1225,height=275)
        self.tvreportInv1=ttk.Treeview(self.frm1, columns =(1,2,3,4,5,6,7,8,9,10,11,12,13,14), show='headings')
        self.tvreportInv1.column(1, width=100,anchor='center')
        self.tvreportInv1.column(2, width=100,anchor='center')
        self.tvreportInv1.column(3, width=80,anchor='center')
        self.tvreportInv1.column(4, width=60,anchor='center')
        self.tvreportInv1.column(5, width=80,anchor='center')
        self.tvreportInv1.column(6, width=80,anchor='center')
        self.tvreportInv1.column(7, width=60,anchor='center')
        self.tvreportInv1.column(8, width=80,anchor='center')
        self.tvreportInv1.column(9, width=80,anchor='center')
        self.tvreportInv1.column(10, width=60,anchor='center')
        self.tvreportInv1.column(11, width=80,anchor='center')
        self.tvreportInv1.column(12, width=80,anchor='center')
        self.tvreportInv1.column(13, width=60,anchor='center')
        self.tvreportInv1.column(14, width=80,anchor='center')
        self.tvreportInv1.heading(1, text ='Codigo')
        self.tvreportInv1.heading(2, text ='Producto')
        self.tvreportInv1.heading(3, text ='Cantidades')
        self.tvreportInv1.heading(4, text ='C/U')
        self.tvreportInv1.heading(5, text ='Total')
        self.tvreportInv1.heading(6, text ='Cantidades')
        self.tvreportInv1.heading(7, text ='C/U')
        self.tvreportInv1.heading(8, text ='Total')
        self.tvreportInv1.heading(9, text ='Cantidades')
        self.tvreportInv1.heading(10, text ='C/U')
        self.tvreportInv1.heading(11, text ='Total')
        self.tvreportInv1.heading(12, text ='Cantidades')
        self.tvreportInv1.heading(13, text ='C/U')
        self.tvreportInv1.heading(14, text ='Total')
        self.style =ttk.Style(self.tvreportInv1)
        self.style.configure('Treeview',  
                              background="#F0E68C",
                              foreground="black",
                              rowheight=15,
                              fieldbackground="white")
        self.style.map("Treeview",background=[('selected', 'green')])
        self.tvreportInv1.pack(padx=10, pady=10, fill='both', expand=True)
        self.scrollbar1 = ttk.Scrollbar(self.labelframe20, orient=tk.VERTICAL, command=self.tvreportInv1.yview)
        self.tvreportInv1.configure(yscroll=self.scrollbar1.set)
        self.scrollbar1.place(x=1220, y=163, height=259)
        self.hscrollbar = ttk.Scrollbar(self.labelframe20, orient=tk.HORIZONTAL)
        self.hscrollbar.config(command=self.tvreportInv1.xview)
        self.hscrollbar.place(x=6,y=420,width=1226,height=15)



    def widgets_reportes_all_inventario2(self):
        self.style = ttk.Style()
        self.style.configure("MyEntry.TEntry",
                        # Fondo rojo.
                        fieldbackground="#FFFFFF",
                        # Color de texto azul.
                        foreground="#0000ff")
        self.style = ttk.Style()
        self.style.theme_use("clam")
        self.style.configure("TNotebook.Tab", background="#FFFFFF", font=font.Font(family="verdana", size=10, weight="bold"))
        self.style =ttk.Style(self.labelframe20)
        self.style.configure('TLabelframe', background='#FFFFFF', weight="bold")
        self.style = ttk.Style()
        self.style.configure("MyEntry.TEntry",
        # Blanco.
        fieldbackground="#FFFFFF",
        # Color de texto azul.
        foreground="#0000ff")
        self.label=ttk.Label(self.labelframe20, text="Inventario Inicial",foreground='blue',background='#ADD8E6',font=font.Font(family="verdana", size=15, weight = "bold"),anchor='center')
        self.label.place(x=238,y=148,width=245,height=30)
        self.label=ttk.Label(self.labelframe20, text="Compras",foreground='blue',background='#6495ED',font=font.Font(family="verdana", size=15, weight = "bold"),anchor='center')
        self.label.place(x=480,y=148,width=245,height=30)
        self.label=ttk.Label(self.labelframe20, text="Salidas",foreground='blue',background='#00BFFF',font=font.Font(family="verdana", size=15, weight = "bold"),anchor='center')
        self.label.place(x=725,y=148,width=245,height=30)
        self.label=ttk.Label(self.labelframe20, text="Inventario Final",foreground='blue',background='#1E90FF',font=font.Font(family="verdana", size=15, weight = "bold"),anchor='center')
        self.label.place(x=970,y=148,width=250,height=30)


    def generar_one_itmes_inventory(self):
        codigo_report = (self.codigo_report.get(),)
        one_itmes = self.base_datos_inventario.reportes_all_inventario1(codigo_report,codigo_report,codigo_report,codigo_report,codigo_report,
                                                                        codigo_report,codigo_report,codigo_report,codigo_report,codigo_report,
                                                                        codigo_report,codigo_report,codigo_report,codigo_report,codigo_report,
                                                                        codigo_report,codigo_report,codigo_report)  
        self.tvreportInv1.delete(*self.tvreportInv1.get_children())
        x = tk.messagebox.askquestion(message="¿Desea generar Reportes de Invenatrios?", title="Rpeortes de Inventrios")
        if x == 'yes':
            for i in one_itmes:
                codigo = (i[0])
                Producto = (i[1])                                            
                Cantidades_iniciales = (i[2])       
                CU_iniciales = (i[3])       
                Total_iniciales = (i[4])       
                Cantidades = (i[5])
                CU = (i[6])
                Total1 = (i[7])
                Cantidades2 = (i[8])
                CU_Salidas = (i[9])
                Total2 = (i[10])
                Inventario_Final = (i[11])
                CU = (i[12])
                Costo_total_Final = (i[13])
                self.tvreportInv1.insert('', 'end', values=(codigo,
                                                            Producto,
                                                            '{:,.2f}'.format(Cantidades_iniciales).replace(",", "@").replace(".", ",").replace("@", "."),
                                                            '{:,.2f}'.format(CU_iniciales).replace(",", "@").replace(".", ",").replace("@", "."),
                                                            '{:,.2f}'.format(Total_iniciales).replace(",", "@").replace(".", ",").replace("@", "."),
                                                            '{:,.2f}'.format(Cantidades).replace(",", "@").replace(".", ",").replace("@", "."),
                                                            '{:,.2f}'.format(CU).replace(",", "@").replace(".", ",").replace("@", "."),
                                                            '{:,.2f}'.format(Total1).replace(",", "@").replace(".", ",").replace("@", "."),
                                                            '{:,.2f}'.format(Cantidades2).replace(",", "@").replace(".", ",").replace("@", "."),
                                                            '{:,.2f}'.format(CU_Salidas).replace(",", "@").replace(".", ",").replace("@", "."),
                                                            '{:,.2f}'.format(Total2).replace(",", "@").replace(".", ",").replace("@", "."),
                                                            '{:,.2f}'.format(Inventario_Final).replace(",", "@").replace(".", ",").replace("@", "."),
                                                            '{:,.2f}'.format(CU).replace(",", "@").replace(".", ",").replace("@", "."),
                                                            '{:,.2f}'.format(Costo_total_Final).replace(",", "@").replace(".", ",").replace("@", ".")))    

    def clear_datos_one_productos1(self):
        self.entrycp.delete(0, 'end')
        for item in self.tvreportInv1.get_children():
            self.tvreportInv1.delete(item)


    def generar_one_itmes_inventory_excel(self,worksheet, side=None, blank=True):
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.merge_cells('B1:O1')
        self.ws['B1'] = 'I N V E N T A R I O S'
        self.ws['B1'].alignment = Alignment(horizontal='center', vertical='center')
        self.ws.merge_cells('B3')
        self.ws['B3'] = 'Codigo'
        self.ws['B3'].alignment = Alignment(horizontal='center', vertical='center')
        self.ws.merge_cells('C3')
        self.ws['C3'] =  'Producto'
        self.ws['C3'].alignment = Alignment(horizontal='center', vertical='center')
        self.ws.merge_cells('D2:F2')
        self.ws['D2'] = 'Inventario Inicial'
        self.ws['D2'].alignment = Alignment(horizontal='center', vertical='center')
        self.ws.merge_cells('G2:I2')
        self.ws['G2'] = 'Compras'
        self.ws['G2'].alignment = Alignment(horizontal='center', vertical='center')
        self.ws.merge_cells('J2:L2')
        self.ws['J2'] = 'Salidas'
        self.ws['J2'].alignment = Alignment(horizontal='center', vertical='center')
        self.ws.merge_cells('M2:O2')
        self.ws['M2'] = 'Inventario Final'
        self.ws['M2'].alignment = Alignment(horizontal='center', vertical='center')
        self.ws['D3'] = 'Cantidades'
        self.ws['D3'].alignment = Alignment(horizontal='center', vertical='center')
        self.ws['E3'] = 'C/U'
        self.ws['E3'].alignment = Alignment(horizontal='center', vertical='center')
        self.ws['F3'] = 'Total'
        self.ws['F3'].alignment = Alignment(horizontal='center', vertical='center')
        self.ws['G3'] = 'Cantidades'
        self.ws['G3'].alignment = Alignment(horizontal='center', vertical='center')
        self.ws['H3'] = 'C/U'
        self.ws['H3'].alignment = Alignment(horizontal='center', vertical='center')
        self.ws['I3'] = 'Total'
        self.ws['I3'].alignment = Alignment(horizontal='center', vertical='center')
        self.ws['J3'] = 'Cantidades'
        self.ws['J3'].alignment = Alignment(horizontal='center', vertical='center')
        self.ws['K3'] = 'C/U'
        self.ws['K3'].alignment = Alignment(horizontal='center', vertical='center')
        self.ws['L3'] = 'Total'
        self.ws['L3'].alignment = Alignment(horizontal='center', vertical='center')
        self.ws['M3'] = 'Cantidades'
        self.ws['M3'].alignment = Alignment(horizontal='center', vertical='center')
        self.ws['N3'] = 'C/U'
        self.ws['N3'].alignment = Alignment(horizontal='center', vertical='center')
        self.ws['O3'] = 'Total'
        self.ws['O3'].alignment = Alignment(horizontal='center', vertical='center')
        for i, item in enumerate(self.tvreportInv1.get_children()):
            values = list(self.tvreportInv1.item(item)['values'])
            self.ws.cell(row=i+4, column=1, value=i+1)
            for j, value in enumerate(values):
                self.ws.cell(row=i+4, column=j+2, value=value)
        messagebox.showinfo(title='Inventarios', message='Descargando Inventarios')
        side = Side(border_style='thin', color='FF0000')
        self.set_border(self.ws, side)
        self.wb.save('Inventarios_one_Item.xlsx')
        self.wb.close()

    #-------------------------------------------INVENTARIO INICIAL --------------------------------------------------#
    def widgets_carga_inventario_inicial(self):
        self.labelframe21=ttk.LabelFrame(self.ventana1, text="Reporte Inventario All Items",borderwidth=20)
        self.labelframe21.grid(column=0, row=0, padx=5, ipadx=716,ipady=450)
        self.style = ttk.Style()
        self.style.configure("MyEntry.TEntry",
                        # Fondo rojo.
                        fieldbackground="#FFFFFF",
                        # Color de texto azul.
                        foreground="#0000ff")
        self.style = ttk.Style()
        self.style.theme_use("clam")
        self.style.configure("TNotebook.Tab", background="#FFFFFF", font=font.Font(family="verdana", size=10, weight="bold"))
        self.style =ttk.Style(self.labelframe21)
        self.style.configure('TLabelframe', background='#FFFFFF', weight="bold")
        self.style = ttk.Style()
        self.style.configure("MyEntry.TEntry",
        # Blanco.
        fieldbackground="#FFFFFF",
        # Color de texto azul.
        foreground="#0000ff")
        self.labelpull=ttk.Label(self.labelframe21, text="IMPORTAR INVENTARIO INICIAL",background='#FFFFFF',font=font.Font(family="verdana", size=12, weight = "bold"))
        self.labelpull.place(x=600, y=-17,height=20)
        self.imgb2 = PhotoImage(file="E:/Contabilidad MySQL/Iconos/carpeta-abierta.png")
        self.boton158 = tkinter.Button(self.labelframe21, image=self.imgb2, text="Importar Archivo", fg='Black', bg='#FFFAF0',compound="left",font=font.Font(family="Verdana", size=8, weight = "bold"),command=lambda:[self.importar_archivo()])
        self.boton158.place(x=-19,y=-19,width=195,height=60)#690
        self.imgb3 = PhotoImage(file="E:/Contabilidad MySQL/Iconos/borrar.png")
        self.boton159 = tkinter.Button(self.labelframe21, image=self.imgb3,text="Borrar", fg='Black', bg='#FFFAF0',compound="left",font=font.Font(family="Verdana", size=8, weight = "bold"),command=lambda:[self.clear_datos_one_inv_ini()])
        self.boton159.place(x=171,y=-19,width=195,height=60)#880
        self.frm1 = ttk.LabelFrame(self.labelframe21)
        self.frm1.place(x=7,y=195,width=1225,height=275)
        self.tvinv_inicial=ttk.Treeview(self.frm1, columns =(1,2,3,4,5), show='headings')
        self.tvinv_inicial.column(1, width=100,anchor='center')
        self.tvinv_inicial.column(2, width=100,anchor='center')
        self.tvinv_inicial.column(3, width=80,anchor='center')
        self.tvinv_inicial.column(4, width=60,anchor='center')
        self.tvinv_inicial.column(5, width=80,anchor='center')
        self.tvinv_inicial.heading(1, text ='Codigo')
        self.tvinv_inicial.heading(2, text ='Producto')
        self.tvinv_inicial.heading(3, text ='Cantidades')
        self.tvinv_inicial.heading(4, text ='C/U')
        self.tvinv_inicial.heading(5, text ='Total')
        self.style =ttk.Style(self.tvinv_inicial)
        self.style.configure('Treeview',  
                              background="#F0E68C",
                              foreground="black",
                              rowheight=15,
                              fieldbackground="white")
        self.style.map("Treeview",background=[('selected', 'green')])
        self.tvinv_inicial.pack(padx=10, pady=10, fill='both', expand=True)
        self.scrollbar15 = ttk.Scrollbar(self.labelframe21, orient=tk.VERTICAL, command=self.tvinv_inicial.yview)
        self.tvinv_inicial.configure(yscroll=self.scrollbar15.set)
        self.scrollbar15.place(x=1220,y=213,height=259)
        self.hscrollbar5 = ttk.Scrollbar(self.labelframe21, orient=tk.HORIZONTAL)
        self.hscrollbar5.config(command=self.tvinv_inicial.xview)
        self.hscrollbar5.place(x=6,y=470,width=1226,height=15)

    def importar_archivo(self):
        import xlrd
        import mysql.connector  #pip install mysql-connector-python
        from tkinter import filedialog
        import pymysql
        from pathlib import Path

        # Open the workbook and define the worksheet
        filename = filedialog.askopenfilename(title="Open a File", filetype=(("xlxs files", ".*xlsx"),("All Files", "*.")))
        file_path = Path(filename)

        xlrd.xlsx.ensure_elementtree_imported(False, None)
        xlrd.xlsx.Element_has_iter = True
        book = xlrd.open_workbook(file_path)
        #book = xlrd.open_workbook(path)
        sheet = book.sheet_by_index(0)

        #book = xlrd.open_workbook("C:/Users/carlo_000/Desktop/output.xlsx")
        #sheet = book.sheet_by_name("source")

        # Establish a MySQL connection
        database = mysql.connector.connect(host='127.0.0.1',port="3306",database ='celtics12', user = 'root',password ='Ca22021956*')

        # Get the cursor, which is used to traverse the database, line by line
        cursor = database.cursor()

        # Create the INSERT INTO sql query
        #query = """INSERT INTO python_connector_tb (name1, age, email) VALUES (%s, %s, %s)"""
        query = """INSERT INTO inv_inicial(codigo,Producto,Cantidades,CU,total) VALUES (%s,%s,%s,%s,%s)"""

        # Create a For loop to iterate through each row in the XLS file, starting at row 2 to skip the headers
        for r in range(1, sheet.nrows):
                codigo = sheet.cell(r, 0).value
                Producto = sheet.cell(r, 1).value
                Cantidades = sheet.cell(r, 2).value
                CU = sheet.cell(r, 3).value
                total = sheet.cell(r, 4).value

                # Assign values from each row
                values = (codigo,Producto,Cantidades,CU,total)
                self.tvinv_inicial.insert('', 'end', values=(values))
                # Execute sql Query
                cursor.execute(query, values)

        # Close the cursor
        cursor.close()

        # Commit the transaction
        database.commit()

        # Close the database connection
        database.close()

        # Print results
        print("")
        print("All Done! Bye, for now.")
        print("")
        columns = str(sheet.ncols)
        rows = str(sheet.nrows)
        #print("I just imported " %2B columns %2B " columns and " %2B rows %2B " rows to MySQL!")



        
        # Imprimir resultado
        messagebox.showinfo(title='importar Archivo', message='Done!')
        print("Done! ")
        columns = str(sheet.ncols)
        rows = str(sheet.nrows)
        messagebox.showinfo(title='importar Archivo', message="Acabo de importar")

    def clear_datos_one_inv_ini(self):
        for item in sself.tvinv_inicial.get_children():
            self.tvinv_inicial.delete(item)

    
    




aplicacion1=Aplicacion()